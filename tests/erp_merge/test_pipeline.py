from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from work_order_process.erp_merge.config import load_config
from work_order_process.erp_merge.mapping import convert_old_to_new_columns
from work_order_process.erp_merge.pipeline import (
    build_standard_sheet,
    merge_erp_sources,
    write_standard_sheet,
)
from work_order_process.erp_schema import standard_headers


def _write_source(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
    metadata_rows: int,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row_number in range(metadata_rows):
        sheet.append([f"说明 {row_number + 1}"])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _write_rules(path: Path, mappings: dict[str, str]) -> None:
    columns = ["新ERP字段", *mappings]
    frame = pd.DataFrame(
        [
            ["旧ERP字段", *mappings.values()],
            ["样例", *("" for _ in mappings)],
        ],
        columns=columns,
    )
    frame.to_excel(path, index=False)


@pytest.fixture
def source_files(tmp_path: Path) -> tuple[Path, Path, Path]:
    new_file = tmp_path / "new_erp_2026-07-15.xlsx"
    old_file = tmp_path / "old_erp_2026-07-14.xlsx"
    rule_file = tmp_path / "rules.xlsx"

    new_headers = [
        "合同编号",
        "销售组织",
        "签约客户",
        "标的行编码",
        "执行明细id",
        "企业版销售合同明细id",
        "营销平台",
        "明细运维开始开始日期",
        "明细运维结束日期",
        "产品金额",
        "合同申请日期",
        "当年应分摊金额",
        "去年同期应分摊金额",
        "备注",
    ]
    _write_source(
        new_file,
        new_headers,
        [
            [
                "NEW-1",
                "新销售组织",
                "新客户",
                "NEW-LINE-1",
                "NEW-EXEC-1",
                "DUP-1",
                "深圳分公司",
                "2026-01-01",
                "2026-12-31",
                "3650",
                "2026-01-02",
                "111",
                "222",
                "保留",
            ],
            ["NEW-TOTAL", "", "", "", "", "", "", "", "", "", "", "", "合计："],
        ],
        metadata_rows=1,
    )

    old_headers = [
        "合同编号",
        "签署公司",
        "合同分录ID",
        "执行明细id",
        "是否标准合同",
        "核算收入类型分组",
        "合同类型",
        "营销平台旧",
        "累计开票金额",
        "累计回款金额",
        "累计收入金额-去年同期",
        "分成比例",
        "运维开始",
        "运维结束",
        "产品金额旧",
        "当年原始分摊",
        "去年原始分摊",
        "汇总标记",
    ]
    _write_source(
        old_file,
        old_headers,
        [
            [
                "OLD-DUP",
                "旧签署公司",
                "DUP-1",
                "DUP-EXEC-1",
                "统签散开合同",
                "软件销售",
                "实施合同",
                "海南分公司",
                "1000",
                "800",
                "600",
                "25%",
                "2025-01-01",
                "2025-12-31",
                "3650",
                "10",
                "20",
                "",
            ],
            [
                "OLD-KEEP",
                "旧签署公司",
                "OLD-LINE-2",
                "OLD-EXEC-2",
                "运维收入暂估合同",
                "运维服务",
                "运维合同",
                "海南分公司",
                "1,000",
                "800",
                "600",
                "25%",
                "2025-01-01",
                "2025-12-31",
                "3650",
                "333",
                "444",
                "",
            ],
            [
                "OLD-KEEP",
                "旧签署公司",
                "OLD-LINE-2",
                "OLD-EXEC-2",
                "运维收入暂估合同",
                "运维服务",
                "运维合同",
                "海南分公司",
                "1,000",
                "800",
                "600",
                "25%",
                "2025-01-01",
                "2025-12-31",
                "3650",
                "333",
                "444",
                "",
            ],
            ["OLD-TOTAL", "", "TOTAL", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "合计:"],
        ],
        metadata_rows=3,
    )

    _write_rules(
        rule_file,
        {
            "序号": "/",
            "合同编号": "合同编号",
            "销售组织": "签署公司",
            "签约客户": "签署公司",
            "企业版销售合同明细id": "合同分录ID",
            "标的行编码": "合同分录ID",
            "执行明细id": "执行明细id",
            "合同类型": "是否标准合同为统签散开合同时转换",
            "暂估运维运营": "是否标准合同为运维收入暂估合同时转换",
            "虚拟合同": "是否标准合同为虚拟销售合同时转换",
            "业务类型": "按核算收入类型分组转换",
            "合同分类": "按合同类型转换",
            "营销平台": "营销平台旧",
            "累计开票金额": "累计开票金额",
            "累计回款金额": "累计回款金额",
            "去年同期收入金额": "累计收入金额-去年同期",
            "明细运维开始开始日期": "运维开始",
            "明细运维结束日期": "运维结束",
            "产品金额": "产品金额旧",
            "当年应分摊金额": "当年原始分摊",
            "去年同期应分摊金额": "去年原始分摊",
        },
    )
    return new_file, old_file, rule_file


def test_merge_sources_detects_headers_drops_totals_deduplicates_and_maps(
    source_files: tuple[Path, Path, Path],
) -> None:
    new_file, old_file, rule_file = source_files
    generated_at = datetime(2026, 7, 16, 9, 8, 7)

    result = merge_erp_sources(new_file, old_file, rule_file, generated_at)

    assert result["合同编号"].tolist() == ["NEW-1", "OLD-KEEP"]
    assert result["数据来源"].tolist() == ["新ERP", "旧ERP"]
    assert result["文件生成时间戳"].tolist() == ["20260716090807"] * 2
    assert result["文件来源时间戳"].tolist() == ["20260715"] * 2
    assert result.loc[1, "标的行编码"] == "OLD-LINE-2"
    assert result.loc[1, "合同类型"] == "普通销售合同"
    assert result.loc[1, "暂估运维运营"] == "是"
    assert result.loc[1, "业务类型"] == "运维服务费"
    assert result.loc[1, "合同分类"] == "运维合同"
    assert result.loc[1, "营销平台"] == "广西分公司"
    assert result.loc[1, "体系工程师"] == "黄迪"
    assert result.loc[1, "累计开票金额"] == pytest.approx(250.0)
    assert result.loc[1, "累计回款金额"] == pytest.approx(200.0)
    assert result.loc[1, "去年同期收入金额"] == pytest.approx(150.0)


def test_old_mapping_keeps_historical_contract_business_and_amount_rules() -> None:
    old = pd.DataFrame(
        {
            "是否标准合同": ["统签散开合同", "虚拟销售合同"],
            "核算收入类型分组": ["软件销售", "未知类型"],
            "合同类型": ["实施合同", "其他合同"],
            "累计开票金额": ["2,000", "100"],
            "分成比例": ["25%", "0.5"],
        }
    )
    output_columns = ["合同类型", "暂估运维运营", "虚拟合同", "业务类型", "合同分类", "累计开票金额"]
    rules = pd.Series(
        {
            "合同类型": "统签散开合同规则",
            "暂估运维运营": "运维收入暂估合同规则",
            "虚拟合同": "虚拟销售合同规则",
            "业务类型": "核算收入类型分组规则",
            "合同分类": "合同类型规则",
            "累计开票金额": "累计开票金额",
        }
    )

    result = convert_old_to_new_columns(old, output_columns, rules, load_config())

    assert result["合同类型"].tolist() == ["统签散开合同", "普通销售合同"]
    assert result["暂估运维运营"].tolist() == ["否", "否"]
    assert result["虚拟合同"].tolist() == ["否", "是"]
    assert result["业务类型"].tolist() == ["软件产品", "其他"]
    assert result["合同分类"].tolist() == ["非运维合同", "/"]
    assert result["累计开票金额"].tolist() == [500.0, 50.0]


def test_build_standard_sheet_preserves_original_amortization_and_appends_calculations(
    source_files: tuple[Path, Path, Path],
) -> None:
    new_file, old_file, rule_file = source_files
    merged = merge_erp_sources(
        new_file,
        old_file,
        rule_file,
        datetime(2026, 7, 16, 9, 8, 7),
    )

    result = build_standard_sheet(
        merged,
        previous_period=("2025-01-01", "2025-12-31"),
        current_period=("2026-01-01", "2026-12-31"),
    )

    assert result.columns.tolist() == standard_headers()
    assert len(result.columns) == 78
    assert result.columns[-9:].tolist() == standard_headers()[-9:]
    assert result.loc[0, "当年应分摊金额"] == pytest.approx(111.0)
    assert result.loc[0, "去年同期应分摊金额"] == pytest.approx(222.0)
    assert result.loc[0, "今年按期分摊服务费"] == pytest.approx(3650.0)
    assert result.loc[1, "去年按期分摊服务费"] == pytest.approx(3650.0)


def test_write_standard_sheet_streams_only_sheet1_with_standard_headers(
    source_files: tuple[Path, Path, Path], tmp_path: Path
) -> None:
    new_file, old_file, rule_file = source_files
    merged = merge_erp_sources(
        new_file,
        old_file,
        rule_file,
        datetime(2026, 7, 16, 9, 8, 7),
    )
    standard = build_standard_sheet(
        merged,
        previous_period=("2025-01-01", "2025-12-31"),
        current_period=("2026-01-01", "2026-12-31"),
    )
    output_file = tmp_path / "standard.xlsx"

    write_standard_sheet(standard, output_file)

    workbook = load_workbook(output_file, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    assert list(next(rows)) == standard_headers()
    first_data_row = dict(zip(standard_headers(), next(rows), strict=True))
    assert first_data_row["合同编号"] == "NEW-1"
    workbook.close()
