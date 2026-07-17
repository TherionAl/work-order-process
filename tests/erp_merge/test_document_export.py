from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from work_order_process import erp_import
from work_order_process.erp_merge.pipeline import write_document_workbook
from work_order_process.erp_schema import standard_headers


def _standard_frame() -> pd.DataFrame:
    values = {header: f"value-{index}" for index, header in enumerate(standard_headers())}
    values.update(
        {
            "合同编号": "CONTRACT-001",
            "明细运维开始开始日期": datetime(2026, 1, 1),
            "明细运维结束日期": datetime(2026, 12, 31),
            "产品金额": 3650.5,
            "当年应分摊金额": 101.25,
            "去年同期应分摊金额": 99.75,
        }
    )
    return pd.DataFrame([values], columns=standard_headers())


def test_write_document_workbook_adds_nonimportable_document_layout(tmp_path: Path) -> None:
    output_file = tmp_path / "erp-document.xlsx"

    write_document_workbook(_standard_frame(), output_file)

    workbook = load_workbook(output_file, data_only=True)
    assert workbook.sheetnames == ["说明", "文档数据"]
    data_sheet = workbook["文档数据"]
    headers = [cell.value for cell in data_sheet[1]]
    values = [cell.value for cell in data_sheet[2]]
    row = dict(zip(headers, values, strict=True))
    assert len(headers) == 79
    assert headers == ["文档类别", *standard_headers()]
    assert row["文档类别"] == "文档行"
    assert row["合同编号"] == "CONTRACT-001"
    assert row["产品金额"] == 3650.5
    assert row["当年应分摊金额"] == 101.25
    assert row["去年同期应分摊金额"] == 99.75
    assert row["明细运维开始开始日期"] == datetime(2026, 1, 1)
    assert row["明细运维结束日期"] == datetime(2026, 12, 31)
    assert data_sheet[2][headers.index("明细运维开始开始日期")].number_format == "yyyy-mm-dd"
    assert data_sheet.freeze_panes == "A2"
    with pytest.raises(ValueError, match="ERP 标准 Sheet1"):
        erp_import.find_standard_sheet(workbook)
    workbook.close()


def test_write_document_workbook_requires_exact_standard_headers(tmp_path: Path) -> None:
    invalid_frame = _standard_frame().loc[:, list(reversed(standard_headers()))]

    with pytest.raises(ValueError):
        write_document_workbook(invalid_frame, tmp_path / "invalid.xlsx")
