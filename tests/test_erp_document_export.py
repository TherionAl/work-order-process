from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from work_order_process.erp_schema import STANDARD_ERP_COLUMN_MAP, standard_headers


class FakeCursor:
    def __init__(self, rows: list[tuple[object, ...]]) -> None:
        self.rows = rows
        self.statements: list[tuple[str, tuple[object, ...] | None]] = []

    def __enter__(self) -> FakeCursor:
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        return None

    def execute(self, statement: str, parameters=None) -> None:
        self.statements.append((statement, parameters))

    def fetchone(self) -> tuple[int]:
        return (len(self.rows),)

    def fetchall(self) -> list[tuple[object, ...]]:
        return []

    def __iter__(self):
        return iter(self.rows)


class FakeConnection:
    def __init__(self, rows: list[tuple[object, ...]]) -> None:
        self.cursor_instance = FakeCursor(rows)
        self.closed = False

    def cursor(self) -> FakeCursor:
        return self.cursor_instance

    def close(self) -> None:
        self.closed = True


def test_export_erp_snapshot_document_reads_the_imported_snapshot(
    monkeypatch, tmp_path: Path
) -> None:
    from work_order_process.erp_document_export import export_erp_snapshot_document

    values = {column: f"value-{index}" for index, (_, column) in enumerate(STANDARD_ERP_COLUMN_MAP)}
    values["contract_id"] = "CONTRACT-SNAPSHOT"
    values["create_date"] = "20260717"
    row = tuple(values[column] for _, column in STANDARD_ERP_COLUMN_MAP)
    connection = FakeConnection([row])
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection, cursors=SimpleNamespace(SSCursor=object)),
    )
    output_file = tmp_path / "erp-document.xlsx"

    report = export_erp_snapshot_document(
        SimpleNamespace(host="fake", port=3306, user="fake", password="fake", database="fake"),
        "20260717",
        output_file,
    )

    assert report == {"create_date": "20260717", "rows": 1, "file": output_file.name}
    assert connection.closed is True
    assert all(parameters == ("20260717",) for _, parameters in connection.cursor_instance.statements)
    workbook = load_workbook(output_file, data_only=True)
    assert workbook.sheetnames == ["文档数据"]
    data_sheet = workbook["文档数据"]
    headers = [cell.value for cell in data_sheet[1]]
    values = [cell.value for cell in data_sheet[2]]
    row = dict(zip(headers, values, strict=True))
    assert headers == ["文档类别", *standard_headers()]
    assert row["合同编号"] == "CONTRACT-SNAPSHOT"
    workbook.close()
