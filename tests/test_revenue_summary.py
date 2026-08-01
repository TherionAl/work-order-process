from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from work_order_process import revenue_summary
from work_order_process.revenue_summary import (
    _METRIC_SQL,
    ENGLISH_HEADERS,
    PERSISTED_COLUMNS,
    build_revenue_rows,
    export_revenue_workbook,
    fetch_revenue_metrics,
    load_revenue_targets,
)


def test_persisted_revenue_generation_does_not_create_schema(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class Cursor:
        def __enter__(self) -> Cursor:
            return self

        def __exit__(self, *_: object) -> None:
            return None

    class Connection:
        def __init__(self) -> None:
            self.commits = 0

        def __enter__(self) -> Connection:
            return self

        def __exit__(self, *_: object) -> None:
            return None

        def cursor(self) -> Cursor:
            return Cursor()

        def commit(self) -> None:
            self.commits += 1

    connection = Connection()
    monkeypatch.setattr(
        revenue_summary,
        "ensure_revenue_summary_schema",
        lambda config: pytest.fail("ordinary revenue persistence must not create schema"),
    )
    monkeypatch.setattr(
        revenue_summary,
        "load_revenue_targets",
        lambda *args: {"platform": Decimal("100")},
    )
    monkeypatch.setattr(revenue_summary, "validate_revenue_snapshot", lambda *args: 1)
    monkeypatch.setattr(
        revenue_summary,
        "fetch_revenue_metrics",
        lambda *args, **kwargs: {"platform": {}},
    )
    monkeypatch.setattr(revenue_summary, "require_revenue_metrics", lambda *args: None)
    monkeypatch.setattr(
        revenue_summary,
        "build_revenue_rows",
        lambda **kwargs: [{"sales_platform": "platform"}],
    )
    monkeypatch.setattr(revenue_summary, "replace_revenue_rows", lambda *args, **kwargs: None)
    monkeypatch.setattr(revenue_summary, "export_revenue_workbook", lambda *args: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    report = revenue_summary.generate_revenue_summary(
        SimpleNamespace(
            host="db",
            port=3306,
            user="user",
            password="secret",
            database="warehouse",
        ),
        target_file=tmp_path / "targets.xlsx",
        year=2026,
        month=7,
        output_dir=tmp_path,
        erp_create_date="20260731",
        output_path=tmp_path / "summary.xlsx",
        persist=True,
    )

    assert report["persisted"] is True
    assert connection.commits == 1


def test_load_revenue_targets_reads_selected_period_and_ignores_total(tmp_path: Path) -> None:
    path = tmp_path / "targets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append([None, "年", "月", "营销平台", "收入目标值"])
    sheet.append([None, "合计", None, None, "=SUM(E4:E5)"])
    sheet.append([None, 2026, 6, "厦门分公司", 1080000])
    sheet.append([None, 2026, 6, "河北分公司", 500000])
    sheet.append([None, 2026, 7, "厦门分公司", 1200000])
    workbook.save(path)

    assert load_revenue_targets(path, year=2026, month=6) == {
        "厦门分公司": Decimal("1080000"),
        "河北分公司": Decimal("500000"),
    }


def test_build_revenue_rows_calculates_four_groups_and_null_zero_denominator() -> None:
    targets = {
        "厦门分公司": Decimal("100"),
        "河北分公司": Decimal("50"),
    }
    metrics = {
        "厦门分公司": {
            "recognized_revenue": Decimal("20"),
            "recognized_revenue_excluding_estimate": Decimal("18"),
            "prior_year_recognized_revenue": Decimal("10"),
            "contracts_on_hand_amount": Decimal("30"),
            "prior_year_contracts_on_hand_amount": Decimal("15"),
            "signing_completed_amount": Decimal("40"),
            "prior_year_signing_amount": Decimal("20"),
        },
        "河北分公司": {
            "recognized_revenue": Decimal("0"),
            "prior_year_recognized_revenue": Decimal("0"),
            "contracts_on_hand_amount": Decimal("5"),
            "prior_year_contracts_on_hand_amount": Decimal("0"),
            "signing_completed_amount": Decimal("0"),
            "prior_year_signing_amount": Decimal("0"),
        },
    }

    rows = build_revenue_rows(
        year=2026,
        month=6,
        erp_create_date="20260717",
        targets=targets,
        metrics=metrics,
    )

    xiamen, hebei = rows
    assert xiamen["recognized_revenue"] == Decimal("20.00")
    assert xiamen["recognized_revenue_excluding_estimate"] == Decimal("18")
    assert xiamen["revenue_completion_rate"] == Decimal("0.200000")
    assert xiamen["contracts_on_hand_yoy_amount"] == Decimal("15.00")
    assert xiamen["contracts_on_hand_yoy_rate"] == Decimal("1.000000")
    assert xiamen["recognized_revenue_yoy_rate"] == Decimal("0.800000")
    assert xiamen["signing_yoy_rate"] == Decimal("1.000000")

    assert hebei["revenue_completion_rate"] == Decimal("0.000000")
    assert hebei["contracts_on_hand_yoy_rate"] is None
    assert hebei["recognized_revenue_yoy_rate"] is None
    assert hebei["signing_yoy_rate"] is None


def test_export_revenue_workbook_writes_english_chinese_and_total_rows(tmp_path: Path) -> None:
    path = tmp_path / "revenue.xlsx"
    rows = build_revenue_rows(
        year=2026,
        month=6,
        erp_create_date="20260717",
        targets={"厦门分公司": Decimal("100")},
        metrics={
            "厦门分公司": {
                "recognized_revenue": Decimal("20"),
                "prior_year_recognized_revenue": Decimal("10"),
                "contracts_on_hand_amount": Decimal("30"),
                "prior_year_contracts_on_hand_amount": Decimal("15"),
                "signing_completed_amount": Decimal("40"),
                "prior_year_signing_amount": Decimal("20"),
            }
        },
    )

    export_revenue_workbook(path, rows)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    assert [sheet.cell(1, column).value for column in range(2, len(ENGLISH_HEADERS) + 2)] == list(
        ENGLISH_HEADERS
    )
    assert sheet.cell(2, 2).value == "统计年"
    assert sheet.cell(3, 2).value == "合计"
    assert sheet.cell(4, 4).value == "厦门分公司"
    assert sheet.cell(3, 5).value == "=SUM(E4:E4)"
    workbook.close()


class _FakeCursor:
    def __init__(
        self,
        rows: list[tuple[object, ...]] | None = None,
        fetchone_values: list[tuple[object, ...]] | None = None,
    ) -> None:
        self.rows = rows or []
        self.fetchone_values = list(fetchone_values or [])
        self.executed: list[tuple[str, object]] = []

    def execute(self, statement: str, parameters=None) -> None:
        self.executed.append((statement, parameters))

    def fetchall(self) -> list[tuple[object, ...]]:
        return self.rows

    def fetchone(self) -> tuple[object, ...] | None:
        if not self.fetchone_values:
            return None
        return self.fetchone_values.pop(0)


def test_fetch_revenue_metrics_uses_confirmed_filters_and_adjusted_amortization() -> None:
    cursor = _FakeCursor(
        [
            (
                "厦门分公司",
                Decimal("20"),
                Decimal("18"),
                Decimal("10"),
                Decimal("30"),
                Decimal("15"),
                Decimal("40"),
                Decimal("20"),
            )
        ]
    )

    metrics = fetch_revenue_metrics(cursor, erp_create_date="20260717", year=2026, month=6)

    statement, parameters = cursor.executed[0]
    assert "other_business_type = '非税票据'" in statement
    assert "is_estimated_ops = '否'" not in statement.split("AS recognized_revenue,")[0]
    assert (
        "is_estimated_ops = '否'"
        in statement.split("AS recognized_revenue,")[1].split(
            "AS recognized_revenue_excluding_estimate,"
        )[0]
    )
    assert "cur_year_adjusted_amort" in statement
    assert "prev_year_adjusted_amort" in statement
    assert parameters[-1] == "20260717"
    assert metrics["厦门分公司"]["contracts_on_hand_amount"] == Decimal("30")
    assert metrics["厦门分公司"]["recognized_revenue_excluding_estimate"] == Decimal("18")


def test_replace_revenue_rows_deletes_month_before_inserting_platforms() -> None:
    cursor = _FakeCursor(fetchone_values=[(1,)])
    rows = build_revenue_rows(
        year=2026,
        month=6,
        erp_create_date="20260717",
        targets={"厦门分公司": Decimal("100")},
        metrics={},
    )

    revenue_summary.replace_revenue_rows(
        cursor,
        year=2026,
        month=6,
        rows=rows,
    )

    delete_statement, delete_parameters = cursor.executed[0]
    assert delete_statement.startswith("DELETE FROM ops_service_revenue_monthly")
    assert delete_parameters == (2026, 6)
    statement, parameters = cursor.executed[1]
    assert "INSERT INTO ops_service_revenue_monthly" in statement
    assert "ON DUPLICATE KEY UPDATE" in statement
    assert parameters[0:3] == (2026, 6, "厦门分公司")
    assert cursor.executed[2][0].startswith("SELECT COUNT(*) FROM ops_service_revenue_monthly")


def test_replace_revenue_rows_rejects_published_row_count_mismatch() -> None:
    cursor = _FakeCursor(fetchone_values=[(0,)])
    rows = build_revenue_rows(
        year=2026,
        month=6,
        erp_create_date="20260717",
        targets={"厦门分公司": Decimal("100")},
        metrics={},
    )

    with pytest.raises(ValueError, match=r"expected=1.*published=0"):
        revenue_summary.replace_revenue_rows(
            cursor,
            year=2026,
            month=6,
            rows=rows,
        )


def test_validate_revenue_snapshot_rejects_missing_snapshot() -> None:
    cursor = _FakeCursor(fetchone_values=[(0, None, None)])

    with pytest.raises(ValueError, match=r"20260799.*不存在"):
        revenue_summary.validate_revenue_snapshot(cursor, "20260799")


def test_validate_revenue_snapshot_rejects_null_allocation_fields() -> None:
    cursor = _FakeCursor(fetchone_values=[(10, 1, 2)])

    with pytest.raises(ValueError, match="年度分摊字段"):
        revenue_summary.validate_revenue_snapshot(cursor, "20260717")


def test_empty_revenue_metrics_are_rejected_before_zero_rows_are_built() -> None:
    with pytest.raises(ValueError, match=r"20260717.*有效营收指标"):
        revenue_summary.require_revenue_metrics({}, "20260717")


def test_cli_generates_revenue_summary_with_explicit_period_and_snapshot(
    monkeypatch, tmp_path: Path
) -> None:
    from work_order_process import cli

    captured: dict[str, object] = {}

    def fake_generate(config, **kwargs):
        captured["config"] = config
        captured.update(kwargs)
        return {
            "stat_year": 2026,
            "stat_month": 6,
            "erp_create_date": "20260717",
            "target_platform_count": 30,
            "metric_platform_count": 30,
            "rows": 30,
            "unmapped_metric_platforms": [],
            "output_path": str(tmp_path / "result.xlsx"),
        }

    monkeypatch.setattr(cli, "generate_revenue_summary", fake_generate, raising=False)
    monkeypatch.setattr(
        cli,
        "load_settings",
        lambda: SimpleNamespace(
            mysql="mysql-config", output_dir=tmp_path, dictionary_path=tmp_path / "dictionary.pdf"
        ),
    )
    monkeypatch.setattr(cli, "assert_schema_current", lambda _: None)
    monkeypatch.setattr(cli.DataDictionary, "from_pdf", lambda _: object())
    monkeypatch.setattr(
        "sys.argv",
        [
            "work_order_process",
            "generate-revenue-summary",
            "--year",
            "2026",
            "--month",
            "6",
            "--revenue-target-file",
            "targets.xlsx",
            "--erp-create-date",
            "20260717",
        ],
    )

    cli.main()

    assert captured["config"] == "mysql-config"
    assert captured["target_file"] == Path("targets.xlsx")
    assert captured["year"] == 2026
    assert captured["month"] == 6
    assert captured["erp_create_date"] == "20260717"


def test_cli_preview_exports_revenue_summary_without_persisting(
    monkeypatch, tmp_path: Path
) -> None:
    from work_order_process import cli

    captured: dict[str, object] = {}

    def fake_generate(config, **kwargs):
        captured["config"] = config
        captured.update(kwargs)
        return {
            "stat_year": 2026,
            "stat_month": 6,
            "erp_create_date": "20260717",
            "target_platform_count": 30,
            "metric_platform_count": 30,
            "rows": 30,
            "unmapped_metric_platforms": [],
            "output_path": str(tmp_path / "preview.xlsx"),
        }

    monkeypatch.setattr(cli, "generate_revenue_summary", fake_generate, raising=False)
    monkeypatch.setattr(
        cli,
        "load_settings",
        lambda: SimpleNamespace(
            mysql="mysql-config", output_dir=tmp_path, dictionary_path=tmp_path / "dictionary.pdf"
        ),
    )
    monkeypatch.setattr(cli, "assert_schema_current", lambda _: None)
    monkeypatch.setattr(cli.DataDictionary, "from_pdf", lambda _: object())
    monkeypatch.setattr(
        "sys.argv",
        [
            "work_order_process",
            "generate-revenue-summary",
            "--year",
            "2026",
            "--month",
            "6",
            "--revenue-target-file",
            "targets.xlsx",
            "--erp-create-date",
            "20260717",
            "--revenue-preview",
        ],
    )

    cli.main()

    assert captured["persist"] is False


def test_revenue_schema_places_erp_snapshot_before_audit_columns_and_rounds_amounts() -> None:
    schema = (
        Path(__file__).resolve().parents[1] / "sql" / "ops_service_revenue_monthly.sql"
    ).read_text(encoding="utf-8")

    assert PERSISTED_COLUMNS[-1] == "erp_create_date"
    assert (
        schema.index("signing_yoy_rate")
        < schema.index("erp_create_date")
        < schema.index("created_at")
    )
    assert "DECIMAL(18,0)" in schema
    assert "ROUND(SUM(CASE" in _METRIC_SQL
    assert "END), 0) AS recognized_revenue" in _METRIC_SQL


def test_revenue_total_view_has_dynamic_total_row_and_sort_order() -> None:
    view_sql = (
        Path(__file__).resolve().parents[1] / "sql" / "v_ops_service_revenue_monthly_with_total.sql"
    ).read_text(encoding="utf-8")

    assert "UNION ALL" in view_sql
    assert "'合计' AS sales_platform" in view_sql
    assert view_sql.count("AS sort_order") == 2
    assert "work_order_process:v0005:revenue_summary_objects" in view_sql
    assert "ROUND(SUM(revenue_target), 0)" in view_sql


def test_compatibility_schema_helper_keeps_v5_view_marker(monkeypatch) -> None:
    marker = "work_order_process:v0005:revenue_summary_objects"

    class Cursor:
        def __init__(self) -> None:
            self.statement = ""
            self.statements: list[str] = []

        def __enter__(self) -> Cursor:
            return self

        def __exit__(self, *_: object) -> None:
            return None

        def execute(self, statement: str) -> None:
            self.statement = statement
            self.statements.append(statement)

        def fetchall(self) -> list[tuple[object, ...]]:
            if self.statement.startswith("SHOW COLUMNS"):
                return [
                    *((column,) for column in PERSISTED_COLUMNS),
                    ("created_at",),
                    ("updated_at",),
                ]
            if "numeric_scale" in self.statement:
                return [(column, 0) for column in revenue_summary.MONEY_COLUMN_COMMENTS]
            return []

    class Connection:
        def __init__(self) -> None:
            self.cursor_instance = Cursor()

        def __enter__(self) -> Connection:
            return self

        def __exit__(self, *_: object) -> None:
            return None

        def cursor(self) -> Cursor:
            return self.cursor_instance

    connection = Connection()
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    revenue_summary.ensure_revenue_summary_schema(
        SimpleNamespace(
            host="db",
            port=3306,
            user="user",
            password="secret",
            database="warehouse",
        )
    )

    assert marker in connection.cursor_instance.statements[-1]


def test_revenue_amounts_round_half_up_to_integer_and_export_without_decimals(
    tmp_path: Path,
) -> None:
    rows = build_revenue_rows(
        year=2026,
        month=6,
        erp_create_date="20260717",
        targets={"厦门分公司": Decimal("100.5")},
        metrics={
            "厦门分公司": {
                "recognized_revenue": Decimal("10.5"),
                "prior_year_recognized_revenue": Decimal("0"),
                "contracts_on_hand_amount": Decimal("20.5"),
                "prior_year_contracts_on_hand_amount": Decimal("0"),
                "signing_completed_amount": Decimal("30.5"),
                "prior_year_signing_amount": Decimal("0"),
            }
        },
    )
    path = tmp_path / "revenue.xlsx"

    export_revenue_workbook(path, rows)

    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    assert rows[0]["revenue_target"] == Decimal("101")
    assert rows[0]["recognized_revenue"] == Decimal("11")
    assert sheet.cell(4, 5).number_format == "#,##0"
    workbook.close()
