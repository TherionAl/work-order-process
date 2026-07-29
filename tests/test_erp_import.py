from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import Workbook

from work_order_process import erp_import
from work_order_process.erp_schema import (
    LEGACY_ERP_COLUMN_MAP,
    STANDARD_ERP_COLUMN_MAP,
    legacy_headers,
    standard_headers,
)


class FakeCursor:
    def __init__(self) -> None:
        self.statements: list[tuple[str, list[object] | tuple[object, ...] | None]] = []
        self.rowcount = 0
        self.stage_rows: list[tuple[object, ...]] = []
        self.last_statement = ""

    def __enter__(self) -> FakeCursor:
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        return None

    def execute(self, statement: str, values=None) -> None:
        self.statements.append((statement, values))
        self.last_statement = statement
        if statement.startswith("INSERT INTO erp_data "):
            self.rowcount = len(self.stage_rows)
        else:
            self.rowcount = 0

    def executemany(self, statement: str, values) -> None:
        rows = list(values)
        self.statements.append((statement, rows))
        self.last_statement = statement
        self.stage_rows.extend(rows)
        self.rowcount = len(rows)

    def fetchall(self) -> list[tuple[object, ...]]:
        return []

    def fetchone(self):
        if "HAVING COUNT(*) > 1" in self.last_statement:
            return None
        if "COUNT(DISTINCT create_date)" in self.last_statement:
            create_date_index = list(erp_import.IMPORT_COLUMNS).index("create_date")
            create_dates = {row[create_date_index] for row in self.stage_rows}
            value = next(iter(create_dates))
            return (len(create_dates), value, value)
        if "COUNT(*) FROM erp_data_import_stage" in self.last_statement:
            return (len(self.stage_rows),)
        if "COUNT(*) FROM erp_data WHERE create_date" in self.last_statement:
            return (len(self.stage_rows),)
        return None


class FakeConnection:
    def __init__(self) -> None:
        self.cursor_instance = FakeCursor()
        self.commit_count = 0
        self.rollback_count = 0
        self.begin_count = 0
        self.closed = False

    def cursor(self) -> FakeCursor:
        return self.cursor_instance

    def commit(self) -> None:
        self.commit_count += 1

    def rollback(self) -> None:
        self.rollback_count += 1

    def begin(self) -> None:
        self.begin_count += 1

    def close(self) -> None:
        self.closed = True


def _fake_config() -> SimpleNamespace:
    return SimpleNamespace(
        host="fake",
        port=3306,
        user="fake",
        password="fake",
        database="fake",
    )


def _standard_frame(**overrides: object) -> pd.DataFrame:
    header_to_column = dict(STANDARD_ERP_COLUMN_MAP)
    values_by_column = {column: None for _, column in STANDARD_ERP_COLUMN_MAP}
    values_by_column.update(
        {
            "contract_id": "CONTRACT-ATOMIC",
            "item_code": "ITEM-ATOMIC",
            "exec_detail_id": "EXEC-ATOMIC",
            "create_date": "20260724120000",
            "contract_days": 365,
            "prev_year_period_start": "2025-01-01",
            "prev_year_period_end": "2025-12-31",
            "prev_year_calc_amort": 10,
            "prev_year_adjusted_amort": 10,
            "cur_year_period_start": "2026-01-01",
            "cur_year_period_end": "2026-12-31",
            "cur_year_calc_amort": 20,
            "cur_year_adjusted_amort": 20,
        }
    )
    values_by_column.update(overrides)
    return pd.DataFrame(
        [[values_by_column[header_to_column[header]] for header in standard_headers()]],
        columns=standard_headers(),
    )


def _workbook_with_sheet(headers: list[str], row: list[object] | None = None) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(headers)
    if row is not None:
        worksheet.append(row)
    return workbook


def _write_workbook(path: Path, headers: list[str], row: list[object]) -> None:
    workbook = _workbook_with_sheet(headers, row)
    workbook.save(path)
    workbook.close()


def _import_values(monkeypatch: pytest.MonkeyPatch, path: Path) -> list[object]:
    connection = FakeConnection()
    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", lambda config: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    erp_import.import_erp_xlsx(
        SimpleNamespace(host="fake", port=3306, user="fake", password="fake", database="fake"),
        path,
    )

    inserts = [
        values
        for statement, values in connection.cursor_instance.statements
        if statement.startswith("INSERT INTO erp_data_import_stage")
    ]
    assert len(inserts) == 1
    return list(inserts[0][0])


def test_find_standard_sheet_uses_headers_when_data_sheet_is_not_first() -> None:
    workbook = Workbook()
    workbook.active.title = "Presentation"
    sheet = workbook.create_sheet("Imported data")
    sheet.append(list(reversed(standard_headers())))

    assert erp_import.find_standard_sheet(workbook) is sheet


@pytest.mark.parametrize(
    "headers",
    [
        legacy_headers()[1:],
        [legacy_headers()[0], legacy_headers()[0], *legacy_headers()[2:]],
        [*standard_headers(), "presentation_only", "presentation_note", "presentation_total"],
    ],
    ids=["missing-header", "duplicate-header", "presentation-layout"],
)
def test_find_standard_sheet_rejects_nonstandard_layouts(headers: list[str]) -> None:
    workbook = _workbook_with_sheet(headers)

    with pytest.raises(ValueError):
        erp_import.find_standard_sheet(workbook)


def test_import_rejects_81_column_presentation_before_database_setup(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "presentation.xlsx"
    _write_workbook(
        path,
        [*standard_headers(), "presentation_only", "presentation_note", "presentation_total"],
        [],
    )
    schema_calls = 0
    connect_calls = 0

    def fail_if_schema_is_called(config) -> None:
        nonlocal schema_calls
        schema_calls += 1

    def fail_if_connect_is_called(**kwargs):
        nonlocal connect_calls
        connect_calls += 1
        raise AssertionError("invalid workbook must not connect to MySQL")

    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", fail_if_schema_is_called)
    monkeypatch.setitem(sys.modules, "pymysql", SimpleNamespace(connect=fail_if_connect_is_called))

    with pytest.raises(ValueError):
        erp_import.import_erp_xlsx(
            SimpleNamespace(host="fake", port=3306, user="fake", password="fake", database="fake"),
            path,
        )

    assert schema_calls == 0
    assert connect_calls == 0


def test_import_legacy_sheet_inserts_null_for_all_allocation_columns(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "legacy.xlsx"
    headers = list(reversed(legacy_headers()))
    values_by_header = {header: f"value-{index}" for index, header in enumerate(headers)}
    for header, column in LEGACY_ERP_COLUMN_MAP:
        if erp_import.CONVERTERS.get(column) is erp_import._to_decimal:
            values_by_header[header] = "1.25"
    values_by_header[legacy_headers()[1]] = "CONTRACT-1"
    _write_workbook(path, headers, [values_by_header[header] for header in headers])

    values = _import_values(monkeypatch, path)
    allocation_columns = [column for _, column in STANDARD_ERP_COLUMN_MAP[-9:]]
    indexes = {column: index for index, (_, column) in enumerate(STANDARD_ERP_COLUMN_MAP)}

    assert [values[indexes[column]] for column in allocation_columns] == [None] * 9


def test_import_reordered_standard_sheet_maps_allocation_values_by_header(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    path = tmp_path / "standard.xlsx"
    headers = list(reversed(standard_headers()))
    values_by_column = {
        column: f"text-{index}" for index, (_, column) in enumerate(STANDARD_ERP_COLUMN_MAP)
    }
    for _, column in STANDARD_ERP_COLUMN_MAP:
        if erp_import.CONVERTERS.get(column) is erp_import._to_decimal:
            values_by_column[column] = "1.25"
    values_by_column.update(
        {
            "contract_id": "CONTRACT-2",
            "cur_year_amort": "101.25",
            "prev_year_amort": "202.50",
            "contract_days": "366",
            "prev_year_period_start": "2025-01-01",
            "prev_year_period_end": "2025-12-31",
            "prev_year_calc_amort": "11.25",
            "prev_year_adjusted_amort": "12.50",
            "cur_year_period_start": "2026-01-01",
            "cur_year_period_end": "2026-12-31",
            "cur_year_calc_amort": "21.25",
            "cur_year_adjusted_amort": "22.50",
        }
    )
    header_to_column = dict(STANDARD_ERP_COLUMN_MAP)
    _write_workbook(
        path,
        headers,
        [values_by_column[header_to_column[header]] for header in headers],
    )

    values = _import_values(monkeypatch, path)
    db_values = dict(zip((column for _, column in STANDARD_ERP_COLUMN_MAP), values, strict=True))

    assert db_values["cur_year_amort"] == 101.25
    assert db_values["prev_year_amort"] == 202.5
    assert db_values["contract_days"] == 366
    assert db_values["prev_year_period_start"] == "2025-01-01"
    assert db_values["prev_year_period_end"] == "2025-12-31"
    assert db_values["prev_year_calc_amort"] == 11.25
    assert db_values["prev_year_adjusted_amort"] == 12.5
    assert db_values["cur_year_period_start"] == "2026-01-01"
    assert db_values["cur_year_period_end"] == "2026-12-31"
    assert db_values["cur_year_calc_amort"] == 21.25
    assert db_values["cur_year_adjusted_amort"] == 22.5


def test_import_standard_dataframe_writes_snapshot_without_intermediate_workbook(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    connection = FakeConnection()
    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", lambda config: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )
    frame = _standard_frame(
        contract_id="CONTRACT-DATAFRAME",
        create_date="20260717123000",
        prev_year_adjusted_amort=12.5,
        cur_year_adjusted_amort=22.5,
    )

    report = erp_import.import_erp_dataframe(
        SimpleNamespace(host="fake", port=3306, user="fake", password="fake", database="fake"),
        frame,
    )

    inserts = [
        values
        for statement, values in connection.cursor_instance.statements
        if statement.startswith("INSERT INTO erp_data_import_stage")
    ]
    assert len(inserts) == 1
    db_values = dict(
        zip(
            (column for _, column in STANDARD_ERP_COLUMN_MAP),
            inserts[0][0],
            strict=True,
        )
    )
    assert db_values["contract_id"] == "CONTRACT-DATAFRAME"
    assert db_values["prev_year_adjusted_amort"] == 12.5
    assert db_values["cur_year_adjusted_amort"] == 22.5
    assert report["create_dates"] == ["20260717"]


def test_date_converter_replaces_old_erp_placeholders_with_null() -> None:
    assert erp_import.convert("archive_date", "/") is None
    assert erp_import.convert("archive_date", "0000-12-30") is None
    assert erp_import.convert("archive_date", "2026-07-13") == "2026-07-13"


def test_import_rejects_incomplete_snapshot_key(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    connection = FakeConnection()
    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", lambda config: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    with pytest.raises(erp_import.ERPImportError, match=r"item_code.*第 2 行"):
        erp_import.import_erp_dataframe(
            _fake_config(),
            _standard_frame(item_code=None),
        )


def test_import_rejects_nonempty_invalid_amount_with_field_and_row(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    connection = FakeConnection()
    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", lambda config: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    with pytest.raises(
        erp_import.ERPImportError,
        match=r"产品金额.*product_amount.*第 2 行.*BAD",
    ):
        erp_import.import_erp_dataframe(
            _fake_config(),
            _standard_frame(product_amount="BAD"),
        )

    assert not any(
        statement.startswith("DELETE FROM erp_data")
        for statement, _ in connection.cursor_instance.statements
    )


def test_import_database_error_does_not_publish_snapshot(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class FailingCursor(FakeCursor):
        def executemany(self, statement: str, values) -> None:
            if statement.startswith("INSERT INTO erp_data_import_stage"):
                raise RuntimeError("simulated database failure")
            super().executemany(statement, values)

    connection = FakeConnection()
    connection.cursor_instance = FailingCursor()
    monkeypatch.setattr(erp_import, "ensure_auxiliary_schema", lambda config: None)
    monkeypatch.setitem(
        sys.modules,
        "pymysql",
        SimpleNamespace(connect=lambda **kwargs: connection),
    )

    with pytest.raises(RuntimeError, match="simulated database failure"):
        erp_import.import_erp_dataframe(
            _fake_config(),
            _standard_frame(),
        )

    assert not any(
        statement.startswith("DELETE FROM erp_data")
        for statement, _ in connection.cursor_instance.statements
    )


def test_publish_staged_snapshot_replaces_exact_date_and_checks_row_count() -> None:
    class PublishCursor(FakeCursor):
        def fetchone(self):
            return (2,)

    cursor = PublishCursor()

    erp_import._publish_staged_snapshot(
        cursor,
        create_date="20260724",
        expected_rows=2,
    )

    statements = [statement for statement, _ in cursor.statements]
    assert statements[0].startswith("DELETE FROM erp_data WHERE create_date")
    assert statements[1].startswith("INSERT INTO erp_data")
    assert "FROM erp_data_import_stage" in statements[1]
    assert statements[2].startswith("SELECT COUNT(*) FROM erp_data")
