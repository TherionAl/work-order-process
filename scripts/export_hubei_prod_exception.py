"""Export filtered Hubei production-exception tickets to an Excel workbook.

The script is read-only for MySQL. It writes an Excel file under ``output/``
unless ``--output-dir`` is supplied.
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import pymysql
import pymysql.cursors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from work_order_process.config import PROJECT_ROOT
from work_order_process.hubei_analysis import (
    AnalysisScope,
    add_scope_arguments,
    build_ticket_version_predicate,
    index_custom_fields_by_ticket_version,
    mysql_config_from_environment,
    scope_from_arguments,
)

MAIN_COLUMNS = [
    ("ticket_id", "工单ID"),
    ("subject", "标题"),
    ("ticket_status", "工单状态"),
    ("ticket_category", "工单类别"),
    ("ticket_type", "工单类型"),
    ("priority_level", "优先级"),
    ("problem_type", "问题类型"),
    ("product_line", "产品线"),
    ("module_name", "模块名称"),
    ("create_dt", "创建时间"),
    ("source_updated_at", "来源更新时间"),
    ("open_dt", "开启时间"),
    ("solve_dt", "解决时间"),
    ("close_dt", "关闭时间"),
    ("wait_dt", "等待时间"),
    ("cust_user_id", "联系人ID"),
    ("cust_user_name", "联系人姓名"),
    ("company_id", "公司ID"),
    ("company_name", "公司名称"),
    ("customer_type", "客户类型"),
    ("customer_industry", "客户行业"),
    ("servicer_user_id", "客服ID"),
    ("servicer_user_name", "客服姓名"),
    ("servicer_group_id", "客服组ID"),
    ("servicer_group_name", "客服组名称"),
    ("creater_id", "创建人ID"),
    ("creater_name", "创建人姓名"),
    ("creater_type", "创建人类型"),
    ("agent_id", "服务商ID"),
    ("ticket_source", "工单来源"),
    ("ticket_template_id", "工单模板ID"),
    ("ticket_template_name", "工单模板名称"),
    ("custom_template_id", "自定义模板ID"),
    ("province", "省份"),
    ("city", "城市"),
    ("district", "区县"),
    ("region_text", "地区原始文本"),
    ("department_id", "内部部门ID"),
    ("department_name", "内部部门名称"),
    ("current_node_name", "当前节点名称"),
    ("current_node_status", "当前节点状态"),
    ("current_node_field", "当前流程节点字段"),
    ("current_node_field_value", "当前流程节点值"),
    ("current_node_started_at", "当前节点进入时间"),
    ("current_node_duration_seconds", "当前节点停留秒数"),
    ("node_field_into_time", "进入节点时间"),
    ("workflow_node_id", "工作流节点ID"),
    ("workflow_id", "工作流ID"),
    ("tag_list", "标签"),
    ("cc_user_id_list", "抄送客服ID列表"),
    ("cc_group_id_list", "抄送客服组ID列表"),
    ("is_deleted", "是否删除"),
    ("deleter_id", "删除人ID"),
    ("delete_dt", "删除时间"),
    ("query_ids", "查询ID集合"),
    ("create_year", "创建年份"),
    ("create_month", "创建月份"),
    ("create_month_label", "创建年月"),
    ("last_sync_at", "最近同步时间"),
    ("sync_status", "同步状态"),
    ("descript", "描述"),
]
EXCLUDED_CUSTOM_FIELDS = {"2023年版服务目录", "2023版服务目录", "[服务目录]（弃用）"}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    add_scope_arguments(parser)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    scope = scope_from_arguments(args)
    main_rows, custom_fields = _read_tickets(scope, limit=args.limit)
    if not main_rows:
        print("未找到符合条件的工单。")
        return

    output_dir = (
        Path(args.output_dir)
        if args.output_dir
        else PROJECT_ROOT / "output" / "hubei_prod_exception"
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"湖北生产环境异常处理工单_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    _write_workbook(output_path, scope, args.limit, main_rows, custom_fields)
    print(f"已导出 {len(main_rows)} 条工单到: {output_path}")


def _read_tickets(
    scope: AnalysisScope, *, limit: int | None
) -> tuple[list[dict[str, Any]], dict[tuple[int, datetime], dict[str, Any]]]:
    mysql = mysql_config_from_environment()
    with pymysql.connect(
        host=mysql.host,
        port=mysql.port,
        user=mysql.user,
        password=mysql.password,
        database=mysql.database,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    ) as connection:
        with connection.cursor() as cursor:
            sql, params = _ticket_query(scope, limit=limit)
            cursor.execute(sql, params)
            main_rows = cursor.fetchall()

        versions = [(row["ticket_id"], row["create_dt"]) for row in main_rows]
        field_rows = _read_custom_fields(connection, versions)
    return main_rows, index_custom_fields_by_ticket_version(field_rows)


def _ticket_query(scope: AnalysisScope, *, limit: int | None) -> tuple[str, tuple[object, ...]]:
    status_placeholders = ", ".join("%s" for _ in scope.statuses)
    sql = f"""
        SELECT t.*
        FROM ticket_detail_main AS t
        WHERE t.province = %s
          AND t.create_dt >= %s
          AND t.create_dt < %s
          AND t.ticket_status IN ({status_placeholders})
          AND EXISTS (
              SELECT 1 FROM ticket_detail_custom_fields AS cf1
              WHERE cf1.ticket_id = t.ticket_id
                AND cf1.create_dt = t.create_dt
                AND cf1.field_name = '【服务目录】'
                AND cf1.field_value LIKE %s
          )
          AND NOT EXISTS (
              SELECT 1 FROM ticket_detail_custom_fields AS cf2
              WHERE cf2.ticket_id = t.ticket_id
                AND cf2.create_dt = t.create_dt
                AND cf2.field_name = %s
                AND cf2.field_value IS NOT NULL
                AND cf2.field_value <> ''
          )
        ORDER BY t.create_dt DESC
    """
    params: tuple[object, ...] = (
        scope.province,
        scope.start,
        scope.end,
        *scope.statuses,
        f"%{scope.service_catalog_contains}%",
        scope.assistance_field,
    )
    if limit is not None:
        sql += " LIMIT %s"
        params += (limit,)
    return sql, params


def _read_custom_fields(
    connection: Any, ticket_versions: list[tuple[int, datetime]]
) -> list[dict[str, Any]]:
    field_rows: list[dict[str, Any]] = []
    for start in range(0, len(ticket_versions), 500):
        predicate, params = build_ticket_version_predicate(ticket_versions[start : start + 500])
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT ticket_id, create_dt, field_order, field_key, field_name,
                       field_value, field_value_json, field_value_type
                FROM ticket_detail_custom_fields
                WHERE {predicate}
                ORDER BY ticket_id, create_dt, field_order
                """,
                params,
            )
            field_rows.extend(cursor.fetchall())
    return field_rows


def _write_workbook(
    output_path: Path,
    scope: AnalysisScope,
    limit: int | None,
    main_rows: list[dict[str, Any]],
    custom_fields: dict[tuple[int, datetime], dict[str, Any]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "工单详情"
    _write_scope_sheet(workbook, scope, limit, len(main_rows))

    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_index, (_column_key, column_name) in enumerate(MAIN_COLUMNS, start=1):
        cell = worksheet.cell(1, column_index, column_name)
        cell.font = header_font
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = header_alignment

    custom_names = sorted(
        {
            field_name
            for fields in custom_fields.values()
            for field_name in fields
            if field_name not in EXCLUDED_CUSTOM_FIELDS
        }
    )
    custom_start_column = len(MAIN_COLUMNS) + 1
    for column_index, field_name in enumerate(custom_names, start=custom_start_column):
        cell = worksheet.cell(1, column_index, f"[自定义] {field_name}")
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = header_alignment

    for row_index, main_row in enumerate(main_rows, start=2):
        for column_index, (column_key, _column_name) in enumerate(MAIN_COLUMNS, start=1):
            cell = worksheet.cell(row_index, column_index, main_row.get(column_key))
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        fields = custom_fields.get((main_row["ticket_id"], main_row["create_dt"]), {})
        for column_index, field_name in enumerate(custom_names, start=custom_start_column):
            worksheet.cell(row_index, column_index, fields.get(field_name))

    for column_index in range(1, worksheet.max_column + 1):
        maximum_length = max(
            (
                len(str(cell.value))
                for cell in worksheet[get_column_letter(column_index)][:100]
                if cell.value
            ),
            default=8,
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(maximum_length + 2, 10), 50
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    workbook.save(output_path)


def _write_scope_sheet(
    workbook: Workbook, scope: AnalysisScope, limit: int | None, matched_count: int
) -> None:
    worksheet = workbook.create_sheet("筛选条件")
    worksheet.append(["参数", "值"])
    worksheet.append(["开始时间（含）", scope.start.strftime("%Y-%m-%d %H:%M:%S")])
    worksheet.append(["结束时间（不含）", scope.end.strftime("%Y-%m-%d %H:%M:%S")])
    worksheet.append(["省份", scope.province])
    worksheet.append(["服务目录包含", scope.service_catalog_contains])
    worksheet.append(["状态", "、".join(scope.statuses)])
    worksheet.append(["结果上限", limit if limit is not None else "未限制"])
    worksheet.append(["导出工单数", matched_count])
    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 32


if __name__ == "__main__":
    main()
