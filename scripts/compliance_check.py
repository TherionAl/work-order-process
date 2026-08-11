"""故障单合规性检查脚本 v4。

合规标准（全部满足即为合规）：
1. [标题] 不为空
2. [自定义] 问题描述 字段的内容包含 操作\\现象\\报错日志 三者任意相关
3. [自定义] 服务完成解决办法 字段的内容包含 故障原因\\处理过程\\脚本\\处理方案附件 任意相关
4. [解决办法] 不能直接只写"已解决"（需有实质内容）
5. [重复检查] 同一天内客户名称+标题+问题描述一致性不过高（非重复工单）
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pymysql
import pymysql.cursors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from work_order_process.config import PROJECT_ROOT
from work_order_process.hubei_analysis import (
    add_scope_arguments,
    mark_duplicate_results,
    mysql_config_from_environment,
    scope_from_arguments,
)

# ---------------------------------------------------------------------------
# 关键词定义
# ---------------------------------------------------------------------------

# 操作/流程相关
OPERATION_KEYWORDS = [
    "操作",
    "流程",
    "步骤",
    "进行",
    "执行",
    "点击",
    "打开",
    "录入",
    "提交",
    "审核",
    "开票",
    "制票",
    "登录",
    "打印",
    "查询",
    "使用",
    "运行",
    "调用",
    "访问",
    "处理",
    "变更",
    "修改",
    "新增",
    "删除",
    "安装",
    "配置",
    "部署",
    "升级",
    "重启",
    "上传",
    "下载",
    "导入",
    "导出",
    "同步",
    "初始化",
    "设置",
    "调整",
]

# 故障现象相关
SYMPTOM_KEYWORDS = [
    "无法",
    "不能",
    "失败",
    "异常",
    "错误",
    "问题",
    "故障",
    "打不开",
    "连不上",
    "超时",
    "崩溃",
    "卡住",
    "乱码",
    "丢失",
    "报错",
    "提示",
    "警告",
    "闪退",
    "无响应",
    "加载",
    "空白",
    "不正确",
    "不一致",
    "缺失",
    "停止",
    "中断",
]

# 报错日志相关
ERROR_LOG_KEYWORDS = [
    "日志",
    "log",
    "error",
    "exception",
    "trace",
    "堆栈",
    "错误码",
    "code",
    "message",
    "failed",
    "null",
    "timeout",
    "java",
    "sql",
    "http",
    "500",
    "404",
    "502",
    "403",
    "401",
    "stack",
    "throw",
    "catch",
    "print",
    "console",
]

# 故障原因相关
CAUSE_KEYWORDS = [
    "原因",
    "导致",
    "因为",
    "由于",
    "引起",
    "造成",
    "引发",
    "配置错误",
    "参数",
    "设置",
    "权限",
    "网络",
    "数据库",
    "服务",
    "版本",
    "不兼容",
    "缺失",
    "过期",
    "冲突",
    "Bug",
    "缺陷",
    "数据",
    "缓存",
    "内存",
    "CPU",
    "磁盘",
    "空间不足",
]

# 处理过程相关
HANDLING_KEYWORDS = [
    "处理",
    "修复",
    "解决",
    "重启",
    "更新",
    "升级",
    "配置",
    "修改",
    "调整",
    "重新",
    "安装",
    "部署",
    "替换",
    "清理",
    "清除",
    "刷新",
    "同步",
    "恢复",
    "回退",
    "关闭",
    "打开",
    "启用",
    "禁用",
    "重置",
    "初始化",
    "卸载",
    "迁移",
]

# 脚本/处理方案附件相关
ATTACHMENT_KEYWORDS = [
    "脚本",
    "附件",
    "方案",
    "文档",
    "sql",
    "SQL",
    "文件",
    "工具",
    "补丁",
    "patch",
    "批处理",
    "命令",
    "代码",
    "程序",
    "软件",
    "工具包",
    "补丁包",
    "升级包",
    "安装包",
    "配置文档",
    "操作手册",
    "指南",
    "说明",
    "截图",
    "视频",
]


def check_operation(text: str) -> tuple[bool, str]:
    """检查是否包含操作/流程相关描述。"""
    hits = [kw for kw in OPERATION_KEYWORDS if kw in text]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_symptom(text: str) -> tuple[bool, str]:
    """检查是否包含故障现象相关描述。"""
    hits = [kw for kw in SYMPTOM_KEYWORDS if kw in text]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_error_log(text: str) -> tuple[bool, str]:
    """检查是否包含报错日志相关描述。"""
    text_lower = text.lower()
    hits = [kw for kw in ERROR_LOG_KEYWORDS if kw.lower() in text_lower]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_cause(text: str) -> tuple[bool, str]:
    """检查是否包含故障原因相关描述。"""
    hits = [kw for kw in CAUSE_KEYWORDS if kw in text]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_handling(text: str) -> tuple[bool, str]:
    """检查是否包含处理过程相关描述。"""
    hits = [kw for kw in HANDLING_KEYWORDS if kw in text]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_attachment(text: str) -> tuple[bool, str]:
    """检查是否包含脚本/附件相关描述。"""
    text_lower = text.lower()
    hits = [kw for kw in ATTACHMENT_KEYWORDS if kw.lower() in text_lower]
    return (len(hits) > 0, ", ".join(hits[:3]))


def check_solution_not_trivial(solution: str) -> tuple[bool, str]:
    """检查解决办法是否不是敷衍的'已解决'。

    返回 (ok, detail)：
    - ok=True：解决办法有实质内容
    - ok=False：解决办法直接写"已解决"或类似敷衍内容
    """
    text = (solution or "").strip()
    if not text:
        return (False, "解决办法为空")
    # 去除空格、标点后比较
    normalized = re.sub(
        r"[\s\-_，。、；：！？\"'（）()【】\[\]{}《》<>·`~!@#$%^&*+=|\\\/?]+", "", text
    )
    trivial_answers = {
        "已解决",
        "已完成",
        "已处理",
        "完成",
        "好了",
        "修复",
        "修复完成",
        "处理完成",
        "已修复",
        "解决",
    }
    if normalized in trivial_answers:
        return (False, f"解决办法敷衍: '{text}'")
    return (True, "")


# ---------------------------------------------------------------------------
# 重复检查
# ---------------------------------------------------------------------------


def _normalize(text: str) -> str:
    """标准化文本：去空格、标点、转小写，用于相似度比较。"""
    text = text.strip().lower()
    text = re.sub(r"[\s\-_，。、；：！？\"'（）()【】\[\]{}《》<>·`~!@#$%^&*+=|\\\/?]+", "", text)
    return text


def _strip_numbers(text: str) -> str:
    """去除文本中的所有数字字符，用于模板化工单的去重判断。"""
    return re.sub(r"[0-9]+", "", text)


def _similarity(a: str, b: str) -> float:
    """计算两个字符串的相似度（0~1），基于 SequenceMatcher（最长公共子序列）。

    相比字符集合 Jaccard，SequenceMatcher 考虑字符顺序，
    避免 "ABCD" 与 "DCBA" 误判为 100% 相似。
    """
    import difflib

    a_norm = _normalize(a)
    b_norm = _normalize(b)
    if not a_norm and not b_norm:
        return 1.0
    if not a_norm or not b_norm:
        return 0.0
    return difflib.SequenceMatcher(None, a_norm, b_norm).ratio()


def _legacy_check_duplicates(results: list[dict], threshold: float = 0.8) -> list[dict]:
    """按「同一天 + 同一客户」分组，仅比较问题描述相似度。

    前提：同一天、同一客户名下，问题描述相似度 >= threshold 才视为重复。
    返回更新了 "is_duplicate" 和 "duplicate_group" 字段的结果列表。
    """
    # 按 (日期, 客户名) 分组
    groups: dict[tuple[str, str], list[tuple[int, dict]]] = defaultdict(list)
    for idx, result in enumerate(results):
        create_dt = result["row"]["create_dt"]
        date_key = create_dt.strftime("%Y-%m-%d") if create_dt else "未知"
        company_key = (result["row"].get("company_name") or "").strip()
        groups[(date_key, company_key)].append((idx, result))

    MIN_DESC_LEN = 10  # 问题描述过短（<10字）时不参与重复判断，避免模板化短文本误判

    for (date_key, _company_key), group in groups.items():
        if len(group) < 2:
            continue
        # 两两比较问题描述
        for i in range(len(group)):
            idx_i, res_i = group[i]
            if res_i.get("is_duplicate"):
                continue
            desc_i = (res_i["row"].get("problem_desc") or "").strip()
            if len(_normalize(desc_i)) < MIN_DESC_LEN:
                continue

            for j in range(i + 1, len(group)):
                idx_j, res_j = group[j]
                if res_j.get("is_duplicate"):
                    continue
                desc_j = (res_j["row"].get("problem_desc") or "").strip()
                if len(_normalize(desc_j)) < MIN_DESC_LEN:
                    continue

                # 去除数字后完全一致 → 仅编号不同的模板工单，不算重复
                if _strip_numbers(desc_i) == _strip_numbers(desc_j):
                    continue

                sim = _similarity(desc_i, desc_j)
                if sim >= threshold:
                    # 标记后一条为重复
                    res_j["is_duplicate"] = True
                    res_j["duplicate_similarity"] = sim
                    res_j["duplicate_group"] = f"{date_key}_{_company_key}_G{i}"
                    res_i.setdefault("duplicate_group", f"{date_key}_{_company_key}_G{i}")

    # 确保每条结果都有字段
    for result in results:
        result.setdefault("is_duplicate", False)
        result.setdefault("duplicate_similarity", 0.0)
        result.setdefault("duplicate_group", "")

    return results


def check_duplicates(results: list[dict], threshold: float = 0.8) -> list[dict]:
    """Compatibility wrapper for the shared documented duplicate rule."""

    return mark_duplicate_results(results, threshold=threshold)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    add_scope_arguments(parser)
    parser.add_argument(
        "--output-file",
        type=Path,
        help="完整的 Excel 输出路径；不能与 --output-dir 同时使用",
    )
    parser.add_argument(
        "--quality-period",
        choices=("manual", "weekly", "monthly"),
        default="manual",
        help="报告执行周期标记；生产调度入口会自动设置",
    )
    args = parser.parse_args(argv)
    if args.output_file is not None and args.output_dir is not None:
        parser.error("--output-file 不能与 --output-dir 同时使用")
    return args


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    scope = scope_from_arguments(args)
    mysql = mysql_config_from_environment()
    conn = pymysql.connect(
        host=mysql.host,
        port=mysql.port,
        user=mysql.user,
        password=mysql.password,
        database=mysql.database,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )

    try:
        with conn.cursor() as cursor:
            status_placeholders = ", ".join("%s" for _ in scope.statuses)
            limit_clause = " LIMIT %s" if args.limit is not None else ""
            cursor.execute(
                f"""
                SELECT t.ticket_id, t.subject, t.descript,
                       t.create_dt, t.solve_dt, t.close_dt,
                       t.servicer_user_name, t.company_name,
                       cf_desc.field_value AS problem_desc,
                       cf_sol.field_value AS solution
                FROM ticket_detail_main AS t
                LEFT JOIN ticket_detail_custom_fields AS cf_desc
                  ON cf_desc.ticket_id = t.ticket_id AND cf_desc.create_dt = t.create_dt
                 AND cf_desc.field_name = '问题描述'
                LEFT JOIN ticket_detail_custom_fields AS cf_sol
                  ON cf_sol.ticket_id = t.ticket_id AND cf_sol.create_dt = t.create_dt
                 AND cf_sol.field_name = '服务完成解决办法'
                WHERE t.province = %s
                  AND t.create_dt >= %s
                  AND t.create_dt < %s
                  AND t.ticket_status IN ({status_placeholders})
                  AND EXISTS (
                      SELECT 1 FROM ticket_detail_custom_fields AS cf1
                      WHERE cf1.ticket_id = t.ticket_id
                        AND cf1.create_dt = t.create_dt
                        AND cf1.field_name = '【服务目录】'
                        AND cf1.field_value LIKE %s
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM ticket_detail_custom_fields AS cf2
                      WHERE cf2.ticket_id = t.ticket_id
                        AND cf2.create_dt = t.create_dt
                        AND cf2.field_name = %s
                        AND cf2.field_value IS NOT NULL
                        AND cf2.field_value <> ''
                  )
                ORDER BY t.create_dt DESC{limit_clause}
                """,
                (
                    scope.province,
                    scope.start,
                    scope.end,
                    *scope.statuses,
                    f"%{scope.service_catalog_contains}%",
                    scope.assistance_field,
                    *((args.limit,) if args.limit is not None else ()),
                ),
            )
            rows = cursor.fetchall()

        print(f"共 {len(rows)} 条工单待检查\n")

        # 逐条检查
        results = []
        compliant_count = 0
        for row in rows:
            subject = (row["subject"] or "").strip()
            problem_desc = (row["problem_desc"] or "").strip()
            solution = (row["solution"] or "").strip()

            # 条件1: 标题不为空
            rule1_ok = len(subject) > 0

            # 条件2: 问题描述包含操作/现象/报错日志任意
            op_ok, op_detail = check_operation(problem_desc)
            sym_ok, sym_detail = check_symptom(problem_desc)
            err_ok, err_detail = check_error_log(problem_desc)
            rule2_ok = op_ok or sym_ok or err_ok
            rule2_detail = []
            if op_ok:
                rule2_detail.append(f"操作:{op_detail}")
            if sym_ok:
                rule2_detail.append(f"现象:{sym_detail}")
            if err_ok:
                rule2_detail.append(f"日志:{err_detail}")
            rule2_detail_str = " | ".join(rule2_detail) if rule2_detail else "无匹配"

            # 条件3: 解决办法包含故障原因/处理过程/脚本附件任意
            cause_ok, cause_detail = check_cause(solution)
            handle_ok, handle_detail = check_handling(solution)
            attach_ok, attach_detail = check_attachment(solution)
            rule3_ok = cause_ok or handle_ok or attach_ok
            rule3_detail = []
            if cause_ok:
                rule3_detail.append(f"原因:{cause_detail}")
            if handle_ok:
                rule3_detail.append(f"过程:{handle_detail}")
            if attach_ok:
                rule3_detail.append(f"脚本:{attach_detail}")
            rule3_detail_str = " | ".join(rule3_detail) if rule3_detail else "无匹配"

            # 条件4: 解决办法不能直接只写"已解决"
            rule4_ok, rule4_detail = check_solution_not_trivial(solution)

            is_compliant = rule1_ok and rule2_ok and rule3_ok and rule4_ok

            results.append(
                {
                    "row": row,
                    "rule1_ok": rule1_ok,
                    "rule2_ok": rule2_ok,
                    "rule2_detail": rule2_detail_str,
                    "rule3_ok": rule3_ok,
                    "rule3_detail": rule3_detail_str,
                    "rule4_ok": rule4_ok,
                    "rule4_detail": rule4_detail,
                    "compliant": is_compliant,
                }
            )

        # 重复检查
        results = check_duplicates(results, threshold=0.8)

        # 最终合规 = 原合规 + 非重复
        for r in results:
            r["final_compliant"] = r["compliant"] and not r["is_duplicate"]

        compliant_count = sum(1 for r in results if r["final_compliant"])
        duplicate_count = sum(1 for r in results if r["is_duplicate"])
        print(f"合规工单数: {compliant_count}/{len(rows)}")
        print(f"重复工单数: {duplicate_count}/{len(rows)}")

    finally:
        conn.close()

    # 统计
    total = len(rows)
    r1_ok = sum(1 for r in results if r["rule1_ok"])
    r2_ok = sum(1 for r in results if r["rule2_ok"])
    r3_ok = sum(1 for r in results if r["rule3_ok"])
    r4_ok = sum(1 for r in results if r["rule4_ok"])
    duplicate_count = sum(1 for r in results if r["is_duplicate"])
    non_compliant = total - compliant_count

    # 不合规原因分布（仅看内容填写类规则1~4，不考虑重复）
    only_r2_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and not r["rule2_ok"]
        and r["rule3_ok"]
        and r["rule4_ok"]
        and not r["is_duplicate"]
    )
    only_r3_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and r["rule2_ok"]
        and not r["rule3_ok"]
        and r["rule4_ok"]
        and not r["is_duplicate"]
    )
    only_r4_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and r["rule2_ok"]
        and r["rule3_ok"]
        and not r["rule4_ok"]
        and not r["is_duplicate"]
    )
    r2_r3_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and not r["rule2_ok"]
        and not r["rule3_ok"]
        and r["rule4_ok"]
        and not r["is_duplicate"]
    )
    r2_r4_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and not r["rule2_ok"]
        and r["rule3_ok"]
        and not r["rule4_ok"]
        and not r["is_duplicate"]
    )
    r3_r4_fail = sum(
        1
        for r in results
        if r["rule1_ok"]
        and r["rule2_ok"]
        and not r["rule3_ok"]
        and not r["rule4_ok"]
        and not r["is_duplicate"]
    )
    multi_fail = sum(
        1
        for r in results
        if not r["final_compliant"]
        and not r["is_duplicate"]
        and sum([not r["rule1_ok"], not r["rule2_ok"], not r["rule3_ok"], not r["rule4_ok"]]) >= 3
    )

    # 输出 Excel
    wb = Workbook()

    # ── Sheet1: 总结 ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "检查总结"

    title_font = Font(bold=True, size=16)
    h2_font = Font(bold=True, size=13)
    ok_font = Font(bold=True, color="375623")
    fail_font = Font(bold=True, color="9C0006")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    row = 1
    ws1.cell(row, 1, "故障单合规性检查报告").font = title_font
    row += 1
    ws1.cell(row, 1, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    row += 1
    ws1.cell(
        row,
        1,
        "时间范围："
        f"{scope.start:%Y-%m-%d %H:%M:%S} 至 {scope.end:%Y-%m-%d %H:%M:%S}（结束不含）"
        f" | 省份：{scope.province} | 服务目录：{scope.service_catalog_contains}"
        " | 排除：已申请总部协作",
    )
    row += 2

    # 总体结论
    ws1.cell(row, 1, "一、总体结论").font = h2_font
    row += 1
    ws1.cell(row, 1, f"检查工单总数：{total}")
    row += 1
    c_cell = ws1.cell(row, 1, f"最终合规工单数：{compliant_count}")
    c_cell.font = ok_font if compliant_count > 0 else fail_font
    row += 1
    nc_cell = ws1.cell(row, 1, f"不合规工单数：{non_compliant}")
    nc_cell.font = fail_font if non_compliant > 0 else ok_font
    row += 1
    rate = compliant_count / total * 100 if total > 0 else 0
    ws1.cell(row, 1, f"合规率：{rate:.1f}%")
    row += 1
    dup_cell = ws1.cell(row, 1, f"重复工单数：{duplicate_count}")
    dup_cell.font = fail_font if duplicate_count > 0 else ok_font
    row += 2

    # 各规则检查结果
    ws1.cell(row, 1, "二、各规则检查结果").font = h2_font
    row += 1

    rule_headers = ["检查规则", "合规数", "不合规数", "合规率"]
    for col_idx, h in enumerate(rule_headers, 1):
        cell = ws1.cell(row, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1

    rules_data = [
        ("规则1：标题非空", r1_ok, total - r1_ok),
        ("规则2：问题描述含操作/现象/报错日志", r2_ok, total - r2_ok),
        ("规则3：解决办法含故障原因/处理过程/脚本附件", r3_ok, total - r3_ok),
        ("规则4：解决办法不能只写'已解决'", r4_ok, total - r4_ok),
        (
            "规则5：非重复工单（同一天客户+标题+描述相似度<80%）",
            total - duplicate_count,
            duplicate_count,
        ),
        ("最终合规（1~5条都满足）", compliant_count, non_compliant),
    ]
    for label, ok_count, fail_count in rules_data:
        pct = ok_count / total * 100 if total > 0 else 0
        ws1.cell(row, 1, label)
        c1 = ws1.cell(row, 2, ok_count)
        c1.alignment = Alignment(horizontal="center")
        c2 = ws1.cell(row, 3, fail_count)
        c2.alignment = Alignment(horizontal="center")
        c3 = ws1.cell(row, 4, f"{pct:.1f}%")
        c3.alignment = Alignment(horizontal="center")
        # 着色
        if pct >= 80:
            c3.fill = ok_fill
        elif pct < 50:
            c3.fill = fail_fill
        row += 1

    row += 1

    # 不合规原因分布
    ws1.cell(row, 1, "三、不合规原因分布").font = h2_font
    row += 1

    reason_headers = ["不合规原因", "工单数", "占比"]
    for col_idx, h in enumerate(reason_headers, 1):
        cell = ws1.cell(row, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1

    reasons = [
        ("仅规则2不满足（问题描述缺失）", only_r2_fail),
        ("仅规则3不满足（解决办法缺失）", only_r3_fail),
        ("仅规则4不满足（解决办法敷衍）", only_r4_fail),
        ("规则2和3都不满足", r2_r3_fail),
        ("规则2和4都不满足", r2_r4_fail),
        ("规则3和4都不满足", r3_r4_fail),
        ("多条规则同时不满足（≥3条）", multi_fail),
    ]
    for label, count in reasons:
        pct = count / total * 100 if total > 0 else 0
        ws1.cell(row, 1, label)
        ws1.cell(row, 2, count).alignment = Alignment(horizontal="center")
        ws1.cell(row, 3, f"{pct:.1f}%").alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # 重复工单明细
    if duplicate_count > 0:
        ws1.cell(row, 1, "四、重复工单明细").font = h2_font
        row += 1
        dup_headers = ["工单ID", "创建时间", "客户", "标题", "相似度", "重复组"]
        for col_idx, h in enumerate(dup_headers, 1):
            cell = ws1.cell(row, col_idx, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for r in results:
            if r["is_duplicate"]:
                ws1.cell(row, 1, r["row"]["ticket_id"])
                ws1.cell(
                    row,
                    2,
                    r["row"]["create_dt"].strftime("%Y-%m-%d") if r["row"]["create_dt"] else "",
                )
                ws1.cell(row, 3, r["row"]["company_name"])
                ws1.cell(row, 4, (r["row"]["subject"] or "")[:50])
                ws1.cell(row, 5, f"{r['duplicate_similarity']:.0%}").alignment = Alignment(
                    horizontal="center"
                )
                ws1.cell(row, 6, r["duplicate_group"])
                row += 1
        row += 1

    # 合规判定标准说明
    ws1.cell(row, 1, "五、合规判定标准").font = h2_font
    row += 1
    standards = [
        "规则1：标题不为空",
        "规则2：自定义字段【问题描述】内容包含 操作/现象/报错日志 任意相关",
        "规则3：自定义字段【服务完成解决办法】内容包含 故障原因/处理过程/脚本附件 任意相关",
        "规则4：解决办法不能直接只写'已解决'（需有实质内容）",
        "规则5：非重复工单 — 同一天内客户名称+标题+问题描述相似度<80%",
        "5条规则全部满足即为合规",
    ]
    for s in standards:
        ws1.cell(row, 1, f"• {s}")
        row += 1

    # Sheet1 列宽
    ws1.column_dimensions["A"].width = 55
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 15
    ws1.column_dimensions["D"].width = 15

    # ── Sheet2: 详情 ──────────────────────────────────────────────
    ws2 = wb.create_sheet("工单详情")

    headers = [
        ("ticket_id", "工单ID"),
        ("create_dt", "创建时间"),
        ("company_name", "客户名称"),
        ("servicer_user_name", "客服"),
        ("subject", "标题"),
        ("rule1", "规则1-标题非空"),
        ("problem_desc", "问题描述"),
        ("rule2", "规则2-描述含操作/现象/日志"),
        ("rule2_detail", "规则2命中关键词"),
        ("solution", "解决办法"),
        ("rule3", "规则3-解决含原因/过程/脚本"),
        ("rule3_detail", "规则3命中关键词"),
        ("rule4", "规则4-非敷衍"),
        ("rule4_detail", "规则4详情"),
        ("is_duplicate", "是否重复"),
        ("final_compliant", "是否合规"),
    ]

    dup_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col_idx, (_, col_name) in enumerate(headers, start=1):
        cell = ws2.cell(1, col_idx, col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, result in enumerate(results, start=2):
        r = result["row"]
        values = [
            r["ticket_id"],
            r["create_dt"].strftime("%Y-%m-%d %H:%M") if r["create_dt"] else "",
            r["company_name"],
            r["servicer_user_name"],
            r["subject"],
            "✓" if result["rule1_ok"] else "✗",
            (r["problem_desc"] or "")[:500],
            "✓" if result["rule2_ok"] else "✗",
            result["rule2_detail"],
            (r["solution"] or "")[:500],
            "✓" if result["rule3_ok"] else "✗",
            result["rule3_detail"],
            "✓" if result["rule4_ok"] else "✗",
            result["rule4_detail"],
            "重复" if result["is_duplicate"] else "",
            "合规" if result["final_compliant"] else "不合规",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws2.cell(row_idx, col_idx, value)
            if col_idx in (6, 8, 11, 13):
                cell.fill = ok_fill if str(value).startswith("✓") else fail_fill
            if col_idx == 15:  # 重复列
                if result["is_duplicate"]:
                    cell.fill = dup_fill
                    cell.font = Font(bold=True, color="9C6500")
            if col_idx == len(headers):
                cell.fill = ok_fill if result["final_compliant"] else fail_fill
                cell.font = Font(bold=True)

    # Sheet2 列宽
    for col_idx in range(1, ws2.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 18
    ws2.column_dimensions["E"].width = 30
    ws2.column_dimensions["G"].width = 40
    ws2.column_dimensions["I"].width = 40
    ws2.column_dimensions["J"].width = 40
    ws2.column_dimensions["L"].width = 40

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(ws2.max_column)}{ws2.max_row}"

    # Sheet3: 回写范围。回写脚本必须读取此页，并且只接受湖北省报告。
    ws3 = wb.create_sheet("回写范围")
    ws3.append(["键", "值"])
    for key, value in (
        ("report_type", "hubei_compliance_check"),
        ("province", scope.province),
        ("quality_period", args.quality_period),
        ("service_catalog_contains", scope.service_catalog_contains),
        ("start", scope.start.isoformat(sep=" ")),
        ("end", scope.end.isoformat(sep=" ")),
    ):
        ws3.append([key, value])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 50

    # 保存
    if args.output_file is not None:
        output_path = args.output_file
    else:
        output_dir = (
            Path(args.output_dir)
            if args.output_dir
            else PROJECT_ROOT / "output" / "compliance_check"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"故障单合规性检查_{timestamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n已输出到: {output_path}")

    # 终端打印摘要
    print("\n=== 统计摘要 ===")
    print(f"规则1-标题非空: {r1_ok}/{total}")
    print(f"规则2-问题描述含操作/现象/日志: {r2_ok}/{total}")
    print(f"规则3-解决办法含原因/过程/脚本: {r3_ok}/{total}")
    print(f"规则5-非重复工单: {total - duplicate_count}/{total}")
    print(f"最终合规: {compliant_count}/{total}")


if __name__ == "__main__":
    main()
