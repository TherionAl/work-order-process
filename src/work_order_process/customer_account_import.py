"""客户台账明细汇总表 Excel → MySQL 导入。"""

from __future__ import annotations

import logging
import math
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .auxiliary_schema import ensure_auxiliary_schema as ensure_auxiliary_schema
from .config import MySQLConfig
from .import_failures import FailureCollector, ImportFailure

logger = logging.getLogger(__name__)

PARSE_FAILURE_STAGE = "parse"
STAGE_FAILURE_STAGE = "customer_account_stage"
PUBLISH_FAILURE_STAGE = "customer_account_publish"
STAGE_FAILURE_MESSAGES = {
    STAGE_FAILURE_STAGE: "customer account staging failed",
    PUBLISH_FAILURE_STAGE: "customer account publishing failed",
}


class CustomerAccountImportError(RuntimeError):
    """Raised when a customer-account snapshot cannot be safely published."""

    def __init__(
        self,
        message: str,
        *,
        failure: ImportFailure | None = None,
    ) -> None:
        super().__init__(message)
        self.failure = failure


# Excel 列名 → DB 列名（按顺序对应 Sheet1 的 40 列）
COLUMN_MAP = [
    ("营销平台", "marketing_platform"),
    ("合同签约客户", "contract_sign_customer"),
    ("最终使用客户", "final_user_customer"),
    ("应收年运维费", "annual_ops_fee"),
    ("业务分类", "business_category"),
    ("是否纳入当年运维收费目标", "is_in_target"),
    ("服务到期时间", "service_expire_date"),
    ("签约进度", "sign_progress"),
    ("合同编码", "contract_code"),
    ("标的行编码", "item_code"),
    ("省份", "province"),
    ("城市", "city"),
    ("区县", "district"),
    ("运维收费项", "ops_item"),
    ("环境项目名称", "env_project_name"),
    ("客户类型", "customer_type"),
    ("销售部门", "sales_dept"),
    ("销售人员", "sales_person"),
    ("当年未签分类", "unsigned_category"),
    ("不纳入原因分类", "exclude_reason_category"),
    ("不纳入原因说明", "exclude_reason_desc"),
    ("合同名称", "contract_name"),
    ("合同申请日期", "contract_apply_date"),
    ("合同类型", "contract_type"),
    ("归档状态", "archive_status"),
    ("是否虚拟合同", "is_virtual"),
    ("运维开始日期", "ops_start_date"),
    ("运维结束日期", "ops_end_date"),
    ("明细合同金额", "detail_amount"),
    ("预计确收金额", "expected_revenue"),
    ("预计回款金额", "expected_collection"),
    ("已确收金额", "actual_revenue"),
    ("已回款金额", "actual_collection"),
    ("实际验收日期", "acceptance_date"),
    ("合同次数", "contract_count"),
    ("付款方式", "payment_method"),
    ("单位联系人", "contact_person"),
    ("单位联系方式", "contact_phone"),
    ("客户沟通情况", "communication_detail"),
    ("备注", "remark"),
]

INSERT_SQL = (
    "INSERT INTO customer_account ("
    + ", ".join(col for _, col in COLUMN_MAP)
    + ", create_date) VALUES ("
    + ", ".join(["%s"] * (len(COLUMN_MAP) + 1))
    + ")"
)

IMPORT_COLUMNS = tuple(column for _, column in COLUMN_MAP) + ("create_date",)
STAGE_TABLE = "customer_account_import_stage"
CREATE_STAGE_SQL = (
    f"CREATE TEMPORARY TABLE {STAGE_TABLE} AS "
    f"SELECT {', '.join(IMPORT_COLUMNS)} FROM customer_account WHERE 1 = 0"
)
STAGE_INSERT_SQL = (
    f"INSERT INTO {STAGE_TABLE} ({', '.join(IMPORT_COLUMNS)}) VALUES "
    f"({', '.join(['%s'] * len(IMPORT_COLUMNS))})"
)
PUBLISH_STAGE_SQL = (
    f"INSERT INTO customer_account ({', '.join(IMPORT_COLUMNS)}) "
    f"SELECT {', '.join(IMPORT_COLUMNS)} FROM {STAGE_TABLE}"
)


def _to_date(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    return s


def _to_decimal(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError, TypeError:
        return None


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except ValueError, TypeError:
        return None


def _to_str(value) -> str | None:
    if value is None or value == "":
        return None
    s = str(value).strip()
    return s if s else None


CONVERTERS = {
    "annual_ops_fee": _to_decimal,
    "service_expire_date": _to_date,
    "contract_apply_date": _to_date,
    "ops_start_date": _to_date,
    "ops_end_date": _to_date,
    "detail_amount": _to_decimal,
    "expected_revenue": _to_decimal,
    "expected_collection": _to_decimal,
    "actual_revenue": _to_decimal,
    "actual_collection": _to_decimal,
    "acceptance_date": _to_date,
    "contract_count": _to_int,
}


def _to_date_strict(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue
    raise ValueError("unsupported date")


def _to_decimal_strict(value: object) -> float | None:
    if value is None or value == "":
        return None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    converted = float(text)
    if not math.isfinite(converted):
        raise ValueError("numeric value must be finite")
    return converted


def _to_int_strict(value: object) -> int | None:
    if value is None or value == "":
        return None
    converted = _to_decimal_strict(value)
    if converted is None:
        return None
    if not converted.is_integer():
        raise ValueError("integer value must be mathematically integral")
    return int(converted)


STRICT_CONVERTERS = {
    **{
        column: _to_decimal_strict
        for column in (
            "annual_ops_fee",
            "detail_amount",
            "expected_revenue",
            "expected_collection",
            "actual_revenue",
            "actual_collection",
        )
    },
    "contract_count": _to_int_strict,
    "service_expire_date": _to_date_strict,
    "contract_apply_date": _to_date_strict,
    "ops_start_date": _to_date_strict,
    "ops_end_date": _to_date_strict,
    "acceptance_date": _to_date_strict,
}


def convert(col_name: str, value) -> object:
    fn = CONVERTERS.get(col_name)
    if fn:
        return fn(value)
    return _to_str(value)


def convert_strict(column: str, value: object, *, source_row: int) -> object:
    """Convert a workbook cell, rejecting malformed nonempty numeric values."""
    converter = STRICT_CONVERTERS.get(column)
    if converter is None:
        return _to_str(value)
    try:
        return converter(value)
    except (TypeError, ValueError) as exc:
        message = f"{column} contains an invalid nonempty value at source row {source_row}"
        failure = FailureCollector(limit=1).capture(
            stage=PARSE_FAILURE_STAGE,
            exc=CustomerAccountImportError(message),
            source_row=source_row,
        )
        raise CustomerAccountImportError(message, failure=failure) from exc


def prepare_customer_account_row(
    values: Sequence[object], *, source_row: int, create_date: str
) -> list[object] | None:
    """Strictly convert one source row or discard it under the name-cleaning rule."""
    prepared = [
        convert_strict(
            column, values[index] if index < len(values) else None, source_row=source_row
        )
        for index, (_, column) in enumerate(COLUMN_MAP)
    ]
    if prepared[1] is None and prepared[2] is None:
        return None
    return [*prepared, create_date]


def _load_stage_rows(
    cursor: Any,
    rows: Iterable[Sequence[object]],
    *,
    create_date: str,
    batch_size: int,
) -> dict[str, int]:
    """Strictly validate workbook rows and bulk-load accepted rows into a temp table."""
    cursor.execute(CREATE_STAGE_SQL)
    counts = {"rows": 0, "accepted": 0, "inserted": 0, "cleaned": 0}
    batch: list[list[object]] = []
    for source_row, values in enumerate(rows, start=2):
        counts["rows"] += 1
        prepared = prepare_customer_account_row(
            values, source_row=source_row, create_date=create_date
        )
        if prepared is None:
            counts["cleaned"] += 1
            continue
        counts["accepted"] += 1
        batch.append(prepared)
        if len(batch) >= batch_size:
            cursor.executemany(STAGE_INSERT_SQL, batch)
            counts["inserted"] += len(batch)
            batch = []
    if batch:
        cursor.executemany(STAGE_INSERT_SQL, batch)
        counts["inserted"] += len(batch)

    cursor.execute(f"SELECT COUNT(*) FROM {STAGE_TABLE}")
    staged_rows = cursor.fetchone()[0]
    if staged_rows != counts["accepted"]:
        raise CustomerAccountImportError(
            f"expected {counts['accepted']} staged rows, found {staged_rows}"
        )
    return counts


def _publish_staged_snapshot(cursor: Any, *, create_date: str, expected_rows: int) -> None:
    """Replace exactly one formal snapshot from the validated temporary stage."""
    cursor.execute("DELETE FROM customer_account WHERE create_date = %s", (create_date,))
    cursor.execute(PUBLISH_STAGE_SQL)
    cursor.execute("SELECT COUNT(*) FROM customer_account WHERE create_date = %s", (create_date,))
    published_rows = cursor.fetchone()[0]
    if published_rows != expected_rows:
        raise CustomerAccountImportError(
            f"expected {expected_rows} rows, published {published_rows}"
        )


def _connect(config: MySQLConfig) -> Any:
    import pymysql

    return pymysql.connect(
        host=config.host,
        port=config.port,
        user=config.user,
        password=config.password,
        database=config.database,
        charset="utf8mb4",
        autocommit=False,
    )


def _import_customer_account_snapshot(
    config: MySQLConfig,
    file_path: Path,
    create_date: str,
    sheet_name: str | None,
    batch_size: int,
) -> dict:
    import time

    started = time.time()
    failures = FailureCollector()
    conn: Any | None = None
    wb: Any | None = None
    current_stage = STAGE_FAILURE_STAGE
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        workbook_rows = iter(ws.iter_rows(values_only=True))
        headers = [str(cell).strip() if cell else "" for cell in next(workbook_rows, ())]
        if headers != [header for header, _ in COLUMN_MAP]:
            logger.warning("customer account headers differ from the configured column order")
        conn = _connect(config)
        with conn.cursor() as cursor:
            counts = _load_stage_rows(
                cursor,
                workbook_rows,
                create_date=create_date,
                batch_size=batch_size,
            )
            current_stage = PUBLISH_FAILURE_STAGE
            conn.begin()
            _publish_staged_snapshot(
                cursor,
                create_date=create_date,
                expected_rows=counts["accepted"],
            )
            conn.commit()
    except CustomerAccountImportError as exc:
        if conn is not None:
            conn.rollback()
        if exc.failure is None:
            exc.failure = failures.capture(
                stage=current_stage,
                exc=exc,
                record_id="customer_account_snapshot",
            )
        logger.error(
            "customer account import failed: %s",
            exc.failure.as_dict(),
        )
        raise
    except Exception as exc:
        if conn is not None:
            conn.rollback()
        safe_message = STAGE_FAILURE_MESSAGES.get(current_stage, "customer account import failed")
        failure = failures.capture(
            stage=current_stage,
            exc=CustomerAccountImportError(safe_message),
            record_id="customer_account_snapshot",
        )
        safe_exception = CustomerAccountImportError(safe_message, failure=failure)
        logger.error(
            "customer account import failed: %s",
            failure.as_dict(),
        )
        raise safe_exception from exc
    finally:
        if conn is not None:
            conn.close()
        if wb is not None:
            wb.close()

    seconds = round(time.time() - started, 1)
    logger.info(
        "customer account import complete: inserted %d, cleaned %d, seconds %s",
        counts["inserted"],
        counts["cleaned"],
        seconds,
    )
    return {
        "file": file_path.name,
        **counts,
        "skipped": 0,
        "failed": failures.total,
        "seconds": seconds,
        "create_date": create_date,
        **failures.as_payload(),
    }


def import_customer_account_xlsx(
    config: MySQLConfig,
    file_path: Path,
    create_date: str,
    sheet_name: str | None = None,
    batch_size: int = 5000,
) -> dict:
    """把客户台账明细 Excel 导入 MySQL。

    返回 {"file": ..., "rows": ..., "inserted": ..., "skipped": ..., "seconds": ...}
    """
    return _import_customer_account_snapshot(config, file_path, create_date, sheet_name, batch_size)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="客户台账明细数据入库")
    parser.add_argument("--file", required=True, help="Excel 文件路径")
    parser.add_argument("--create-date", required=True, help="数据日期，如 20260710")
    parser.add_argument("--sheet", default=None, help="Sheet 名称（默认第一个）")
    parser.add_argument("--batch-size", type=int, default=5000, help="每批提交行数")
    args = parser.parse_args()

    from .config import load_settings

    settings = load_settings()
    result = import_customer_account_xlsx(
        settings.mysql, Path(args.file), args.create_date, args.sheet, args.batch_size
    )
    print(result)


if __name__ == "__main__":
    main()
