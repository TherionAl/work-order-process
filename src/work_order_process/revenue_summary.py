"""Build and export monthly operations-service revenue summaries."""

from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .config import MySQLConfig, PROJECT_ROOT


ENGLISH_HEADERS = (
    "stat_year",
    "stat_month",
    "sales_platform",
    "revenue_target",
    "recognized_revenue",
    "revenue_completion_rate",
    "contracts_on_hand_amount",
    "prior_year_contracts_on_hand_amount",
    "contracts_on_hand_yoy_amount",
    "contracts_on_hand_yoy_rate",
    "recognized_revenue_excluding_estimate",
    "prior_year_recognized_revenue",
    "recognized_revenue_yoy_amount",
    "recognized_revenue_yoy_rate",
    "signing_completed_amount",
    "prior_year_signing_amount",
    "signing_yoy_amount",
    "signing_yoy_rate",
)

CHINESE_HEADERS = (
    "统计年",
    "统计月",
    "营销平台",
    "收入目标值",
    "确收完成值",
    "收入完成率",
    "在手合同额",
    "去年同期在手合同额",
    "同比增长值",
    "同比增长率",
    "不含暂估确收值",
    "去年同期确收值",
    "确收同比增长值",
    "确收同比增长率",
    "签约完成值",
    "去年同期签约值",
    "签约同比增长值",
    "签约同比增长率",
)

AMOUNT_COLUMNS = {
    "revenue_target",
    "recognized_revenue",
    "contracts_on_hand_amount",
    "prior_year_contracts_on_hand_amount",
    "contracts_on_hand_yoy_amount",
    "recognized_revenue_excluding_estimate",
    "prior_year_recognized_revenue",
    "recognized_revenue_yoy_amount",
    "signing_completed_amount",
    "prior_year_signing_amount",
    "signing_yoy_amount",
}
RATE_COLUMNS = {
    "revenue_completion_rate",
    "contracts_on_hand_yoy_rate",
    "recognized_revenue_yoy_rate",
    "signing_yoy_rate",
}

MONEY_COLUMN_COMMENTS = {
    "revenue_target": "收入目标值（元）",
    "recognized_revenue": "确收完成值（元）",
    "contracts_on_hand_amount": "在手合同额（元）",
    "prior_year_contracts_on_hand_amount": "去年同期在手合同额（元）",
    "contracts_on_hand_yoy_amount": "在手合同同比增长值（元）",
    "recognized_revenue_excluding_estimate": "不含暂估确收值（元）",
    "prior_year_recognized_revenue": "去年同期确收值（元）",
    "recognized_revenue_yoy_amount": "确收同比增长值（元）",
    "signing_completed_amount": "签约完成值（元）",
    "prior_year_signing_amount": "去年同期签约值（元）",
    "signing_yoy_amount": "签约同比增长值（元）",
}

PERSISTED_COLUMNS = (
    *ENGLISH_HEADERS,
    "erp_create_date",
)

UPSERT_SQL = (
    "INSERT INTO ops_service_revenue_monthly ("
    + ", ".join(PERSISTED_COLUMNS)
    + ") VALUES ("
    + ", ".join(["%s"] * len(PERSISTED_COLUMNS))
    + ") ON DUPLICATE KEY UPDATE "
    + ", ".join(
        f"{column} = VALUES({column})"
        for column in PERSISTED_COLUMNS
        if column not in {"stat_year", "stat_month", "sales_platform"}
    )
)

_BASE_ELIGIBLE_CONTRACT = """
    is_public_cloud = '否'
    AND contract_category = '运维合同'
    AND other_business_type = '非税票据'
    AND invalid_contract_type = '有效'
"""

_NO_ESTIMATE_ELIGIBLE_CONTRACT = _BASE_ELIGIBLE_CONTRACT + " AND is_estimated_ops = '否'"

_METRIC_SQL = f"""
SELECT
    sales_platform,
    ROUND(SUM(CASE WHEN {_BASE_ELIGIBLE_CONTRACT} THEN COALESCE(cur_year_revenue, 0) ELSE 0 END), 0) AS recognized_revenue,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} THEN COALESCE(cur_year_revenue, 0) ELSE 0 END), 0) AS recognized_revenue_excluding_estimate,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} THEN COALESCE(prev_year_revenue, 0) ELSE 0 END), 0) AS prior_year_recognized_revenue,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} AND contract_apply_date < %s THEN COALESCE(cur_year_adjusted_amort, 0) ELSE 0 END), 0) AS contracts_on_hand_amount,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} AND contract_apply_date < %s THEN COALESCE(prev_year_adjusted_amort, 0) ELSE 0 END), 0) AS prior_year_contracts_on_hand_amount,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} AND contract_apply_date >= %s AND contract_apply_date < %s THEN COALESCE(product_amount, 0) ELSE 0 END), 0) AS signing_completed_amount,
    ROUND(SUM(CASE WHEN {_NO_ESTIMATE_ELIGIBLE_CONTRACT} AND contract_apply_date >= %s AND contract_apply_date < %s THEN COALESCE(product_amount, 0) ELSE 0 END), 0) AS prior_year_signing_amount
FROM erp_data
WHERE create_date = %s
  AND sales_platform IS NOT NULL
  AND TRIM(sales_platform) <> ''
GROUP BY sales_platform
"""


def load_revenue_targets(path: Path, year: int, month: int) -> dict[str, Decimal]:
    """Read one period's fixed marketing-platform targets from the supplied template."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_row_index, indexes = _find_target_header(worksheet)
        targets: dict[str, Decimal] = {}
        for values in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            platform = _text_at(values, indexes["营销平台"])
            if not platform or platform == "合计":
                continue
            if _int_at(values, indexes["年"]) != year or _int_at(values, indexes["月"]) != month:
                continue
            target = _decimal_at(values, indexes["收入目标值"])
            if target is None:
                raise ValueError(f"营销平台 {platform} 的收入目标值为空。")
            if platform in targets:
                raise ValueError(f"营销平台 {platform} 在 {year}-{month:02d} 存在重复目标值。")
            targets[platform] = _amount(target)
    finally:
        workbook.close()

    if not targets:
        raise ValueError(f"目标文件中不存在 {year}-{month:02d} 的营销平台收入目标值。")
    return targets


def build_revenue_rows(
    *,
    year: int,
    month: int,
    erp_create_date: str,
    targets: Mapping[str, Decimal],
    metrics: Mapping[str, Mapping[str, Decimal | None]],
) -> list[dict[str, Decimal | int | str | None]]:
    """Combine fixed targets with ERP aggregations into database-ready summary rows."""

    rows: list[dict[str, Decimal | int | str | None]] = []
    for platform, target in targets.items():
        platform_metrics = metrics.get(platform, {})
        recognized_revenue = _amount(platform_metrics.get("recognized_revenue"))
        prior_year_recognized_revenue = _amount(platform_metrics.get("prior_year_recognized_revenue"))
        contracts_on_hand_amount = _amount(platform_metrics.get("contracts_on_hand_amount"))
        prior_year_contracts_on_hand_amount = _amount(platform_metrics.get("prior_year_contracts_on_hand_amount"))
        signing_completed_amount = _amount(platform_metrics.get("signing_completed_amount"))
        prior_year_signing_amount = _amount(platform_metrics.get("prior_year_signing_amount"))
        recognized_revenue_excluding_estimate = _amount(
            platform_metrics.get("recognized_revenue_excluding_estimate", recognized_revenue)
        )

        rows.append(
            {
                "stat_year": year,
                "stat_month": month,
                "sales_platform": platform,
                "erp_create_date": erp_create_date,
                "revenue_target": _amount(target),
                "recognized_revenue": recognized_revenue,
                "revenue_completion_rate": _rate(recognized_revenue, _amount(target)),
                "contracts_on_hand_amount": contracts_on_hand_amount,
                "prior_year_contracts_on_hand_amount": prior_year_contracts_on_hand_amount,
                "contracts_on_hand_yoy_amount": _amount(
                    contracts_on_hand_amount - prior_year_contracts_on_hand_amount
                ),
                "contracts_on_hand_yoy_rate": _growth_rate(
                    contracts_on_hand_amount, prior_year_contracts_on_hand_amount
                ),
                "recognized_revenue_excluding_estimate": recognized_revenue_excluding_estimate,
                "prior_year_recognized_revenue": prior_year_recognized_revenue,
                "recognized_revenue_yoy_amount": _amount(
                    recognized_revenue_excluding_estimate - prior_year_recognized_revenue
                ),
                "recognized_revenue_yoy_rate": _growth_rate(
                    recognized_revenue_excluding_estimate, prior_year_recognized_revenue
                ),
                "signing_completed_amount": signing_completed_amount,
                "prior_year_signing_amount": prior_year_signing_amount,
                "signing_yoy_amount": _amount(signing_completed_amount - prior_year_signing_amount),
                "signing_yoy_rate": _growth_rate(signing_completed_amount, prior_year_signing_amount),
            }
        )
    return rows


def export_revenue_workbook(path: Path, rows: list[Mapping[str, Decimal | int | str | None]]) -> None:
    """Export the monthly result in the supplied template's bilingual-header layout."""

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    for column, value in enumerate(ENGLISH_HEADERS, start=2):
        worksheet.cell(1, column, value)
    for column, value in enumerate(CHINESE_HEADERS, start=2):
        worksheet.cell(2, column, value)

    worksheet.cell(3, 2, "合计")
    _write_total_formulas(worksheet, len(rows))

    for row_index, row in enumerate(rows, start=4):
        for column_index, column_name in enumerate(ENGLISH_HEADERS, start=2):
            worksheet.cell(row_index, column_index, row.get(column_name))

    for row in worksheet.iter_rows(min_row=1, max_row=2, min_col=2, max_col=len(ENGLISH_HEADERS) + 1):
        for cell in row:
            cell.font = Font(bold=True)
    for column in range(2, len(ENGLISH_HEADERS) + 2):
        worksheet.column_dimensions[worksheet.cell(1, column).column_letter].width = 20
    for row in range(3, len(rows) + 4):
        for column_index, column_name in enumerate(ENGLISH_HEADERS, start=2):
            if column_name in AMOUNT_COLUMNS:
                worksheet.cell(row, column_index).number_format = '#,##0'
            elif column_name in RATE_COLUMNS:
                worksheet.cell(row, column_index).number_format = '0.00%'

    workbook.save(path)
    workbook.close()


def ensure_revenue_summary_schema(config: MySQLConfig) -> None:
    """Create the small monthly summary table when it does not exist."""

    import pymysql

    table_statement = (PROJECT_ROOT / "sql" / "ops_service_revenue_monthly.sql").read_text(encoding="utf-8")
    view_statement = (PROJECT_ROOT / "sql" / "v_ops_service_revenue_monthly_with_total.sql").read_text(encoding="utf-8")
    with pymysql.connect(
        host=config.host,
        port=config.port,
        user=config.user,
        password=config.password,
        database=config.database,
        charset="utf8mb4",
        autocommit=True,
    ) as connection:
        with connection.cursor() as cursor:
            cursor.execute(table_statement)
            cursor.execute("SHOW COLUMNS FROM ops_service_revenue_monthly")
            column_names = [row[0] for row in cursor.fetchall()]
            cursor.execute(
                "SELECT column_name, numeric_scale FROM information_schema.columns "
                "WHERE table_schema = DATABASE() AND table_name = 'ops_service_revenue_monthly'"
            )
            scales = {row[0]: row[1] for row in cursor.fetchall()}
            money_columns_to_migrate = [
                column for column in MONEY_COLUMN_COMMENTS if scales.get(column) != 0
            ]
            if money_columns_to_migrate:
                cursor.execute(
                    "UPDATE ops_service_revenue_monthly SET "
                    + ", ".join(f"{column} = ROUND({column}, 0)" for column in money_columns_to_migrate)
                )
                cursor.execute(
                    "ALTER TABLE ops_service_revenue_monthly "
                    + ", ".join(
                        f"MODIFY COLUMN {column} DECIMAL(18,0) NOT NULL COMMENT '{MONEY_COLUMN_COMMENTS[column]}'"
                        for column in money_columns_to_migrate
                    )
                )
            if column_names[-3:] != ["erp_create_date", "created_at", "updated_at"]:
                cursor.execute(
                    "ALTER TABLE ops_service_revenue_monthly "
                    "MODIFY COLUMN erp_create_date VARCHAR(8) NOT NULL COMMENT 'ERP快照日期' "
                    "AFTER signing_yoy_rate"
                )
            cursor.execute(view_statement)


def fetch_revenue_metrics(
    cursor: Any,
    *,
    erp_create_date: str,
    year: int,
    month: int,
) -> dict[str, dict[str, Decimal]]:
    """Aggregate the confirmed four revenue groups from one ERP snapshot."""

    current_period_start = date(year, 1, 1)
    current_period_end = _month_end(year, month)
    prior_period_start = date(year - 1, 1, 1)
    prior_period_end = _month_end(year - 1, month)
    cursor.execute(
        _METRIC_SQL,
        (
            current_period_end,
            prior_period_end,
            current_period_start,
            current_period_end,
            prior_period_start,
            prior_period_end,
            erp_create_date,
        ),
    )
    fields = (
        "recognized_revenue",
        "recognized_revenue_excluding_estimate",
        "prior_year_recognized_revenue",
        "contracts_on_hand_amount",
        "prior_year_contracts_on_hand_amount",
        "signing_completed_amount",
        "prior_year_signing_amount",
    )
    metrics: dict[str, dict[str, Decimal]] = {}
    for row in cursor.fetchall():
        platform = str(row[0]).strip()
        metrics[platform] = {
            field: _decimal_value(value)
            for field, value in zip(fields, row[1:], strict=True)
        }
    return metrics


def save_revenue_rows(cursor: Any, rows: list[Mapping[str, Decimal | int | str | None]]) -> None:
    """Upsert one row per statistics month and marketing platform."""

    for row in rows:
        cursor.execute(UPSERT_SQL, tuple(row[column] for column in PERSISTED_COLUMNS))


def generate_revenue_summary(
    config: MySQLConfig,
    *,
    target_file: Path,
    year: int,
    month: int,
    output_dir: Path,
    erp_create_date: str | None = None,
    output_path: Path | None = None,
    persist: bool = True,
) -> dict[str, object]:
    """Load fixed targets, aggregate ERP data, optionally upsert the monthly table, and export Excel."""

    import pymysql

    if month < 1 or month > 12:
        raise ValueError("统计月必须在 1 至 12 之间。")

    targets = load_revenue_targets(target_file, year, month)
    if persist:
        ensure_revenue_summary_schema(config)
    with pymysql.connect(
        host=config.host,
        port=config.port,
        user=config.user,
        password=config.password,
        database=config.database,
        charset="utf8mb4",
        autocommit=False,
    ) as connection:
        with connection.cursor() as cursor:
            if erp_create_date is None:
                cursor.execute("SELECT MAX(create_date) FROM erp_data")
                result = cursor.fetchone()
                erp_create_date = str(result[0] or "").strip() if result else ""
            if not erp_create_date:
                raise ValueError("erp_data 中不存在可用的 ERP 快照。")
            metrics = fetch_revenue_metrics(
                cursor,
                erp_create_date=erp_create_date,
                year=year,
                month=month,
            )
            rows = build_revenue_rows(
                year=year,
                month=month,
                erp_create_date=erp_create_date,
                targets=targets,
                metrics=metrics,
            )
            if persist:
                save_revenue_rows(cursor, rows)
        if persist:
            connection.commit()

    export_path = output_path or output_dir / "revenue_summary" / f"运维服务营收数据表_{year}{month:02d}.xlsx"
    export_revenue_workbook(export_path, rows)
    return {
        "stat_year": year,
        "stat_month": month,
        "erp_create_date": erp_create_date,
        "target_platform_count": len(targets),
        "metric_platform_count": len(metrics),
        "rows": len(rows),
        "unmapped_metric_platforms": sorted(set(metrics) - set(targets)),
        "persisted": persist,
        "output_path": str(export_path),
    }


def _find_target_header(worksheet) -> tuple[int, dict[str, int]]:
    required = ("年", "月", "营销平台", "收入目标值")
    for row_index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        indexes = {str(value).strip(): index for index, value in enumerate(values) if value is not None}
        if all(name in indexes for name in required):
            return row_index, indexes
    raise ValueError("目标文件缺少 年、月、营销平台、收入目标值 列。")


def _write_total_formulas(worksheet, row_count: int) -> None:
    end_row = row_count + 3
    amount_columns = (5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18)
    for column in amount_columns:
        letter = worksheet.cell(1, column).column_letter
        worksheet.cell(3, column, f"=SUM({letter}4:{letter}{end_row})")

    worksheet.cell(3, 7, "=IFERROR(F3/E3,\"\")")
    worksheet.cell(3, 11, "=IFERROR(H3/I3-1,\"\")")
    worksheet.cell(3, 15, "=IFERROR(L3/M3-1,\"\")")
    worksheet.cell(3, 19, "=IFERROR(P3/Q3-1,\"\")")


def _text_at(values: tuple[object, ...], index: int) -> str | None:
    if index >= len(values) or values[index] is None:
        return None
    text = str(values[index]).strip()
    return text or None


def _int_at(values: tuple[object, ...], index: int) -> int | None:
    value = _decimal_at(values, index)
    return int(value) if value is not None else None


def _decimal_at(values: tuple[object, ...], index: int) -> Decimal | None:
    if index >= len(values) or values[index] is None:
        return None
    try:
        return Decimal(str(values[index]).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _amount(value: Decimal | int | str | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _rate(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == 0:
        return None
    return (numerator / denominator).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)


def _growth_rate(current: Decimal, prior: Decimal) -> Decimal | None:
    if prior == 0:
        return None
    return ((current / prior) - Decimal("1")).quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)


def _month_end(year: int, month: int) -> date:
    if month < 1 or month > 12:
        raise ValueError("统计月必须在 1 至 12 之间。")
    if month == 12:
        return date(year + 1, 1, 1)
    return date(year, month + 1, 1)


def _decimal_value(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal("0")
