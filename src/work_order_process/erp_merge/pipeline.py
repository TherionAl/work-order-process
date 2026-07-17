from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

from work_order_process.erp_schema import standard_headers

from .calculator import add_statistical_allocation_columns
from .config import load_config
from .mapping import (
    DATE_FIELDS,
    MONEY_PATTERN,
    add_engineer_column,
    align_new_data,
    convert_old_to_new_columns,
    format_date_fields,
    format_numeric_fields,
    format_text_fields,
    normalize_money_columns,
    normalize_text,
)


MAX_HEADER_SCAN_ROWS = 10
NEW_REQUIRED_COLUMNS = ["合同编号", "销售组织", "签约客户", "标的行编码"]
OLD_REQUIRED_COLUMNS = ["合同编号", "签署公司", "合同分录ID", "是否标准合同"]
SOURCE_COLUMN = "数据来源"
GENERATED_AT_COLUMN = "文件生成时间戳"
SOURCE_DATE_COLUMN = "文件来源时间戳"
EXTRA_OLD_TEXT_COLUMNS = ["其他业务类型", "无效合同类型"]
SNAPSHOT_KEY_COLUMNS = ["合同编号", "标的行编码", "执行明细id"]
DOCUMENT_CATEGORY_HEADER = "文档类别"
DOCUMENT_CATEGORY_VALUE = "文档行"
DOCUMENT_SHEET_NAME = "文档数据"
DOCUMENT_DATE_FIELDS = frozenset(
    [
        *DATE_FIELDS,
        "去年统计起始日期",
        "去年统计截止日期",
        "今年统计起始日期",
        "今年统计截止日期",
    ]
)


def _clean_header(value: object) -> str:
    if value is None or value is pd.NA or value is pd.NaT:
        return ""
    return str(value).strip()


def _find_header_row(raw: pd.DataFrame, required_columns: Iterable[str]) -> int:
    required = set(required_columns)
    for row_index in range(min(MAX_HEADER_SCAN_ROWS, len(raw))):
        row_values = {_clean_header(value) for value in raw.iloc[row_index].tolist()}
        if required.issubset(row_values):
            return row_index
    raise ValueError(f"无法定位包含字段 {sorted(required)} 的表头行，请检查源文件格式。")


def _is_total_cell(value: object) -> bool:
    if value is None or value is pd.NA or value is pd.NaT:
        return False
    return str(value).strip().rstrip("：:") == "合计"


def _drop_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    total_mask = df.map(_is_total_cell).any(axis=1)
    return df.loc[~total_mask].reset_index(drop=True)


def _read_source(file_path: Path, required_columns: Iterable[str]) -> pd.DataFrame:
    raw = pd.read_excel(
        file_path, header=None, dtype=str, keep_default_na=False
    )
    header_row = _find_header_row(raw, required_columns)
    columns = [_clean_header(value) for value in raw.iloc[header_row].tolist()]
    data = raw.iloc[header_row + 1 :].copy()
    data.columns = columns
    data = data.loc[:, [column != "" for column in data.columns]]
    data = data.replace("", pd.NA).dropna(how="all").fillna("")
    return _drop_total_rows(data)


def _load_rules(rule_file: Path) -> tuple[list[str], pd.Series]:
    rules = pd.read_excel(rule_file, dtype=str, keep_default_na=False)
    if rules.shape[1] < 2 or rules.empty:
        raise ValueError("字段对照关系文件缺少目标字段或映射规则。")
    output_columns = [_clean_header(column) for column in rules.columns[1:]]
    mapping = rules.iloc[0, 1:].copy()
    mapping.index = output_columns

    if "体系工程师" not in output_columns:
        if "营销平台" in output_columns:
            output_columns.insert(output_columns.index("营销平台") + 1, "体系工程师")
        else:
            output_columns.append("体系工程师")
        mapping.loc["体系工程师"] = ""
    for column in EXTRA_OLD_TEXT_COLUMNS:
        if column not in output_columns:
            output_columns.append(column)
            mapping.loc[column] = column
    return output_columns, mapping


def _remove_old_rows_existing_in_new(
    old_df: pd.DataFrame, new_df: pd.DataFrame
) -> pd.DataFrame:
    old_key = "合同分录ID"
    new_key = "企业版销售合同明细id"
    if old_key not in old_df.columns:
        raise KeyError(f'旧 ERP 数据源缺少"{old_key}"字段，无法按明细行比对。')
    if new_key not in new_df.columns:
        raise KeyError(f'新 ERP 主数据缺少"{new_key}"字段，无法按明细行比对。')
    new_ids = set(normalize_text(new_df[new_key]))
    new_ids.discard("")
    old_ids = normalize_text(old_df[old_key])
    return old_df.loc[~old_ids.isin(new_ids)].reset_index(drop=True)


def _deduplicate_snapshot_lines(df: pd.DataFrame) -> pd.DataFrame:
    """Match the ERP snapshot unique key before calculating allocation amounts."""
    if not set(SNAPSHOT_KEY_COLUMNS).issubset(df.columns):
        return df

    normalized_keys = pd.DataFrame(
        {
            column: normalize_text(df[column])
            for column in SNAPSHOT_KEY_COLUMNS
        },
        index=df.index,
    )
    complete_key = normalized_keys.ne("").all(axis=1)
    duplicate = complete_key & normalized_keys.duplicated(keep="last")
    return df.loc[~duplicate].reset_index(drop=True)


def _source_date(files: Iterable[Path]) -> str:
    latest = ""
    for file_path in files:
        match = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})", file_path.name)
        if match:
            year, month, day = match.groups()
            latest = max(latest, f"{int(year):04d}{int(month):02d}{int(day):02d}")
    return latest


def merge_erp_sources(
    new_file: Path,
    old_file: Path,
    rule_file: Path,
    generated_at: datetime,
) -> pd.DataFrame:
    config = load_config()
    output_columns, mapping = _load_rules(rule_file)
    new_df = _read_source(new_file, NEW_REQUIRED_COLUMNS)
    old_df = _read_source(old_file, OLD_REQUIRED_COLUMNS)
    old_df = _remove_old_rows_existing_in_new(old_df, new_df)

    new_aligned = align_new_data(new_df, output_columns)
    old_converted = convert_old_to_new_columns(
        old_df, output_columns, mapping, config
    )
    new_aligned[SOURCE_COLUMN] = "新ERP"
    old_converted[SOURCE_COLUMN] = "旧ERP"
    merged = pd.concat([new_aligned, old_converted], ignore_index=True)
    merged = add_engineer_column(merged, config)
    merged = normalize_money_columns(merged)
    merged = format_date_fields(merged)
    merged = format_numeric_fields(merged)
    merged = format_text_fields(merged)
    merged = _deduplicate_snapshot_lines(merged)

    if "合同编号" in merged.columns:
        contract_numbers = normalize_text(merged["合同编号"])
        merged = merged.loc[~contract_numbers.isin({"", "/"})].reset_index(drop=True)
    if "序号" in merged.columns:
        merged["序号"] = range(1, len(merged) + 1)

    for column, default in (("无效合同类型", "有效"), ("其他业务类型", "非税票据")):
        if column in merged.columns:
            merged[column] = normalize_text(merged[column]).replace("", default)

    merged[GENERATED_AT_COLUMN] = generated_at.strftime("%Y%m%d%H%M%S")
    merged[SOURCE_DATE_COLUMN] = _source_date([new_file, old_file])
    middle = [
        column
        for column in merged.columns
        if column not in (SOURCE_COLUMN, GENERATED_AT_COLUMN, SOURCE_DATE_COLUMN)
    ]
    return merged[
        middle + [SOURCE_COLUMN, GENERATED_AT_COLUMN, SOURCE_DATE_COLUMN]
    ]


def build_standard_sheet(
    merged: pd.DataFrame,
    previous_period: tuple[str, str],
    current_period: tuple[str, str],
) -> pd.DataFrame:
    result = merged.copy()
    if "合同申请年份" not in result.columns and "合同申请日期" in result.columns:
        apply_dates = pd.to_datetime(
            normalize_text(result["合同申请日期"]), errors="coerce"
        )
        result["合同申请年份"] = apply_dates.dt.year

    result = add_statistical_allocation_columns(
        result,
        load_config(),
        previous_period[0],
        previous_period[1],
        current_period[0],
        current_period[1],
    )
    result["去年倒签调整后分摊服务费"] = result.pop(
        "去年按期分摊服务费（去掉今年倒签的）"
    )
    result["今年倒签调整后分摊服务费"] = result.pop(
        "今年按期分摊服务费（加上倒签去年的服务费）"
    )

    headers = standard_headers()
    for column in headers:
        if column not in result.columns:
            result[column] = 0.0 if MONEY_PATTERN.search(column) else ""
    return result.loc[:, headers]


def _excel_value(value: object) -> object:
    if value is None or value is pd.NA or value is pd.NaT:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    item = getattr(value, "item", None)
    return item() if callable(item) else value


def write_standard_sheet(df: pd.DataFrame, output_file: Path) -> None:
    headers = standard_headers()
    if df.columns.tolist() != headers:
        raise ValueError("标准 Sheet1 列头或顺序与 78 列契约不一致。")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Sheet1")
    sheet.append(headers)
    for row in df.itertuples(index=False, name=None):
        sheet.append([_excel_value(value) for value in row])
    workbook.save(output_file)


def write_document_rows(
    headers: list[str], rows: Iterable[tuple[object, ...]], output_file: Path
) -> None:
    """Write a readable ERP document workbook from a stream of standard rows."""
    if headers != standard_headers():
        raise ValueError("Document rows must use the complete standard ERP header order.")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)

    sheet = workbook.create_sheet(DOCUMENT_SHEET_NAME)
    sheet.freeze_panes = "A2"
    sheet.append(
        [_document_header_cell(sheet, header) for header in [DOCUMENT_CATEGORY_HEADER, *headers]]
    )
    for row in rows:
        document_row: list[object] = [DOCUMENT_CATEGORY_VALUE]
        for header, value in zip(headers, row, strict=True):
            cell_value = _document_data_value(header, value)
            if header in DOCUMENT_DATE_FIELDS and isinstance(
                cell_value, (datetime, date)
            ):
                cell = WriteOnlyCell(sheet, value=cell_value)
                cell.number_format = "yyyy-mm-dd"
                document_row.append(cell)
            else:
                document_row.append(cell_value)
        sheet.append(document_row)
    workbook.save(output_file)


def _document_date_value(value: object) -> object:
    excel_value = _excel_value(value)
    if excel_value is None or isinstance(excel_value, (datetime, date)):
        return excel_value
    if not isinstance(excel_value, str) or not excel_value.strip():
        return excel_value
    parsed = pd.to_datetime(excel_value.strip(), errors="coerce")
    if pd.isna(parsed) or not 1 <= parsed.year <= 9999:
        return excel_value
    return parsed.to_pydatetime()


def _document_header_cell(sheet, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _document_data_value(header: str, value: object) -> object:
    return (
        _document_date_value(value)
        if header in DOCUMENT_DATE_FIELDS
        else _excel_value(value)
    )


def write_document_workbook(df: pd.DataFrame, output_file: Path) -> None:
    """Write a readable, non-importable ERP document workbook from a frame."""
    headers = standard_headers()
    if df.columns.tolist() != headers:
        raise ValueError("文档工作簿仅接受列头及顺序均符合 78 列标准 Sheet1 的数据。")
    write_document_rows(headers, df.itertuples(index=False, name=None), output_file)
