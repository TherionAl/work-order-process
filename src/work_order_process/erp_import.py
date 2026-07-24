"""ERP 新旧合并数据 Excel → MySQL 导入。"""

from __future__ import annotations

import argparse
import logging
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from .config import MySQLConfig
from .auxiliary_schema import ensure_auxiliary_schema
from .erp_schema import (
    LEGACY_ERP_COLUMN_MAP,
    STANDARD_ERP_COLUMN_MAP,
    legacy_headers,
    standard_headers,
)

logger = logging.getLogger(__name__)

BASELINE_SALES_PLATFORM_CREATE_DATE = "20260713"
SALES_PLATFORM_BASELINE_KEY_COLUMNS = ("contract_id", "item_code", "exec_detail_id")
# Backward-compatible public alias for the historical 69-column contract.
COLUMN_MAP = LEGACY_ERP_COLUMN_MAP
IMPORT_COLUMN_MAP = STANDARD_ERP_COLUMN_MAP
IMPORT_COLUMNS = tuple(column for _, column in IMPORT_COLUMN_MAP)
IMPORT_COLUMN_LABELS = {
    column: header for header, column in IMPORT_COLUMN_MAP
}
STAGE_TABLE = "erp_data_import_stage"
SNAPSHOT_KEY_COLUMNS = ("contract_id", "item_code", "exec_detail_id")
ALLOCATION_COLUMNS = tuple(column for _, column in STANDARD_ERP_COLUMN_MAP[-9:])

STAGE_CREATE_SQL = (
    f"CREATE TEMPORARY TABLE {STAGE_TABLE} ENGINE=InnoDB AS "
    f"SELECT {', '.join(IMPORT_COLUMNS)} FROM erp_data WHERE 1 = 0"
)
STAGE_INSERT_SQL = (
    f"INSERT INTO {STAGE_TABLE} ({', '.join(IMPORT_COLUMNS)}) VALUES ("
    + ", ".join(["%s"] * len(IMPORT_COLUMNS))
    + ")"
)
PUBLISH_INSERT_SQL = (
    f"INSERT INTO erp_data ({', '.join(IMPORT_COLUMNS)}) "
    f"SELECT {', '.join(IMPORT_COLUMNS)} FROM {STAGE_TABLE} "
    "WHERE create_date = %s"
)


class ERPImportError(RuntimeError):
    """Raised when an ERP snapshot cannot be validated or published."""


def _to_date(value) -> str | None:
    """把 Excel 日期值转为 YYYY-MM-DD 字符串。"""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s in {"/", "-", "0", "0.0", "0000-00-00", "0000-12-30"}:
        return None
    try:
        parsed = (
            datetime.strptime(s, "%Y%m%d")
            if len(s) == 8 and s.isdigit()
            else datetime.fromisoformat(s.replace("/", "-"))
        )
    except ValueError:
        return None
    return parsed.strftime("%Y-%m-%d")


def _to_decimal(value) -> Decimal | None:
    """把数值转为 Decimal。"""
    if value is None or value == "":
        return None
    text = str(value).replace(",", "").strip()
    if text in {"", "/"} or text.lower() in {"nan", "nat", "<na>"}:
        return None
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return None
    return result if result.is_finite() else None


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _to_str(value) -> str | None:
    if value is None or value == "":
        return None
    s = str(value).strip()
    return s if s else None


def _to_create_date(value) -> str | None:
    """把 Excel file_timestamp 20260713165231 转为 20260713。"""
    if value is None or value == "":
        return None
    s = str(value).strip()
    return s[:8] if len(s) >= 8 else s


# 值转换函数
CONVERTERS = {
    "seq_no": _to_int,
    "create_date": _to_create_date,
    "contract_apply_date": _to_date,
    "archive_date": _to_date,
    "ops_start_date": _to_date,
    "ops_end_date": _to_date,
    "total_amount": _to_decimal,
    "free_ops_months": _to_int,
    "annual_ops_amount": _to_decimal,
    "detail_qty": _to_int,
    "unit_price": _to_decimal,
    "detail_amount_with_tax": _to_decimal,
    "product_amount": _to_decimal,
    "cum_billing": _to_decimal,
    "cum_collection": _to_decimal,
    "cum_revenue": _to_decimal,
    "cur_year_billing": _to_decimal,
    "prev_year_billing": _to_decimal,
    "cur_year_collection": _to_decimal,
    "prev_year_collection": _to_decimal,
    "cur_year_revenue": _to_decimal,
    "prev_year_revenue": _to_decimal,
    "cur_year_amort": _to_decimal,
    "prev_year_amort": _to_decimal,
    "contract_days": _to_int,
    "prev_year_period_start": _to_date,
    "prev_year_period_end": _to_date,
    "prev_year_calc_amort": _to_decimal,
    "prev_year_adjusted_amort": _to_decimal,
    "cur_year_period_start": _to_date,
    "cur_year_period_end": _to_date,
    "cur_year_calc_amort": _to_decimal,
    "cur_year_adjusted_amort": _to_decimal,
}


def convert(
    col_name: str,
    value,
    *,
    source_row: int | None = None,
    source_name: str | None = None,
) -> object:
    fn = CONVERTERS.get(col_name)
    if fn:
        converted = fn(value)
        if (
            fn is _to_decimal
            and converted is None
            and value is not None
            and str(value).strip() not in {"", "/"}
            and str(value).strip().lower() not in {"nan", "nat", "<na>"}
        ):
            location = (
                f"{source_name} 第 {source_row} 行"
                if source_name and source_row is not None
                else f"第 {source_row} 行"
                if source_row is not None
                else "ERP 数据"
            )
            label = IMPORT_COLUMN_LABELS.get(col_name, col_name)
            raise ERPImportError(
                f"{label}（{col_name}）在 {location}无法解析为金额: {value!r}"
            )
        return converted
    return _to_str(value)


def _header_labels(worksheet: Any) -> list[str]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value).strip() if value is not None else "" for value in first_row]


def find_standard_sheet(workbook) -> Any:
    """Return the only worksheet with exactly the legacy or standard headers.

    The imported data sheet is identified by its complete first-row header set,
    not its sheet name or its position in the workbook.
    """
    accepted_layouts = {frozenset(legacy_headers()), frozenset(standard_headers())}
    matches: list[Any] = []
    duplicate_sheet_names: list[str] = []

    for worksheet in workbook.worksheets:
        headers = _header_labels(worksheet)
        if len(headers) != len(set(headers)):
            duplicate_sheet_names.append(worksheet.title)
            continue
        if frozenset(headers) in accepted_layouts and len(headers) in {69, 78}:
            matches.append(worksheet)

    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError("工作簿中存在多个符合 ERP 标准列头的工作表，无法确定导入目标。")
    if duplicate_sheet_names:
        raise ValueError(
            "未找到 ERP 标准 Sheet1：工作表列头存在重复值（"
            + ", ".join(duplicate_sheet_names)
            + "）。"
        )
    raise ValueError(
        "未找到 ERP 标准 Sheet1：首行必须恰好包含 69 列历史标准列头或 78 列标准列头。"
    )


SalesPlatformBaseline = dict[tuple[str, str | None, str | None], str | None]
SystemEngineerMapping = dict[str, str]


def _baseline_key(row: dict[str, Any]) -> tuple[str, str | None, str | None] | None:
    contract_id = row.get("contract_id")
    if contract_id is None:
        return None
    return (
        str(contract_id),
        row.get("item_code"),
        row.get("exec_detail_id"),
    )


def apply_baseline_sales_platform(
    row: dict[str, Any],
    baseline: SalesPlatformBaseline,
    baseline_create_date: str = BASELINE_SALES_PLATFORM_CREATE_DATE,
) -> bool:
    """Reuse the 20260713 sales_platform for business lines that already existed.

    Only sales_platform is special-cased. All other values in row remain the
    current Excel import values.
    """
    if row.get("create_date") == baseline_create_date:
        return False
    key = _baseline_key(row)
    if key is None or key not in baseline:
        return False
    row["sales_platform"] = baseline[key]
    return True


def load_sales_platform_baseline(cursor: Any, create_date: str = BASELINE_SALES_PLATFORM_CREATE_DATE) -> SalesPlatformBaseline:
    """Load the sales_platform baseline keyed by contract/item/exec detail."""
    cursor.execute(
        """
        SELECT contract_id, item_code, exec_detail_id, sales_platform
        FROM erp_data
        WHERE create_date = %s
        """,
        (create_date,),
    )
    baseline: SalesPlatformBaseline = {}
    for contract_id, item_code, exec_detail_id, sales_platform in cursor.fetchall():
        baseline[(str(contract_id), item_code, exec_detail_id)] = sales_platform
    return baseline


def apply_sales_platform_system_engineer(
    row: dict[str, Any],
    mapping: SystemEngineerMapping | None = None,
) -> bool:
    """Set system_engineer from the configured final sales_platform mapping."""
    if mapping is None:
        from .erp_merge.config import load_config

        mapping = load_config()["体系工程师"]
    sales_platform = row.get("sales_platform")
    if not sales_platform or sales_platform not in mapping:
        return False
    row["system_engineer"] = mapping[sales_platform]
    return True


def _validate_erp_row(
    row: Mapping[str, object],
    source_row: int,
    *,
    require_allocation_fields: bool,
) -> None:
    for column in (*SNAPSHOT_KEY_COLUMNS, "create_date"):
        if row.get(column) is None:
            raise ERPImportError(f"{column} 为空，源数据第 {source_row} 行无法导入。")
    if require_allocation_fields:
        missing = [column for column in ALLOCATION_COLUMNS if row.get(column) is None]
        if missing:
            raise ERPImportError(
                f"标准 ERP 数据第 {source_row} 行缺少年度分摊字段: "
                + ", ".join(missing)
            )


def _insert_stage_batch(
    cursor: Any,
    batch: list[tuple[int, tuple[object, ...]]],
) -> None:
    values = [row_values for _, row_values in batch]
    try:
        cursor.executemany(STAGE_INSERT_SQL, values)
    except Exception as exc:
        first_row = batch[0][0]
        last_row = batch[-1][0]
        raise ERPImportError(
            f"ERP 临时表写入失败，源数据行 {first_row}-{last_row}: {exc}"
        ) from exc


def _validate_staged_snapshot(
    cursor: Any,
    *,
    create_date: str,
    expected_rows: int,
) -> None:
    cursor.execute(f"SELECT COUNT(*) FROM {STAGE_TABLE}")
    staged_rows = int(cursor.fetchone()[0])
    if staged_rows != expected_rows:
        raise ERPImportError(
            f"ERP 临时表行数不一致: expected={expected_rows}, staged={staged_rows}"
        )

    cursor.execute(
        f"""
        SELECT contract_id, item_code, exec_detail_id, COUNT(*)
        FROM {STAGE_TABLE}
        GROUP BY contract_id, item_code, exec_detail_id
        HAVING COUNT(*) > 1
        LIMIT 1
        """
    )
    duplicate = cursor.fetchone()
    if duplicate:
        raise ERPImportError(
            "ERP 快照存在重复业务键: "
            f"contract_id={duplicate[0]}, item_code={duplicate[1]}, "
            f"exec_detail_id={duplicate[2]}, count={duplicate[3]}"
        )

    cursor.execute(
        f"SELECT COUNT(DISTINCT create_date), MIN(create_date), MAX(create_date) "
        f"FROM {STAGE_TABLE}"
    )
    distinct_count, minimum, maximum = cursor.fetchone()
    if int(distinct_count) != 1 or str(minimum) != create_date or str(maximum) != create_date:
        raise ERPImportError(
            "本次 ERP 数据必须且只能包含一个快照日期: "
            f"expected={create_date}, min={minimum}, max={maximum}"
        )


def _publish_staged_snapshot(
    cursor: Any,
    *,
    create_date: str,
    expected_rows: int,
) -> int:
    cursor.execute(
        "DELETE FROM erp_data WHERE create_date = %s",
        (create_date,),
    )
    replaced_rows = max(int(cursor.rowcount), 0)
    cursor.execute(PUBLISH_INSERT_SQL, (create_date,))
    cursor.execute(
        "SELECT COUNT(*) FROM erp_data WHERE create_date = %s",
        (create_date,),
    )
    published_rows = int(cursor.fetchone()[0])
    if published_rows != expected_rows:
        raise ERPImportError(
            f"ERP 正式快照行数不一致: expected={expected_rows}, "
            f"published={published_rows}"
        )
    return replaced_rows


def _import_erp_records(
    config: MySQLConfig,
    excel_rows: Iterable[Mapping[str, object]],
    source_name: str,
    batch_size: int = 5000,
    *,
    require_allocation_fields: bool = True,
) -> dict:
    """Import normalized ERP records using the same rules for every source."""
    import pymysql
    import time

    ensure_auxiliary_schema(config)
    conn = pymysql.connect(
        host=config.host,
        port=config.port,
        user=config.user,
        password=config.password,
        database=config.database,
        charset="utf8mb4",
        autocommit=False,
    )

    reused_baseline_sales_platform = 0
    new_sales_platform = 0
    applied_system_engineer_mapping = 0
    kept_excel_system_engineer = 0
    data_rows = 0
    create_date: str | None = None
    replaced_rows = 0
    started = time.time()

    try:
        with conn.cursor() as cursor:
            sales_platform_baseline = load_sales_platform_baseline(cursor)
            from .erp_merge.config import load_config

            system_engineer_mapping = load_config()["体系工程师"]
            logger.info(
                "已加载 %s 行 %s 营销平台基准",
                len(sales_platform_baseline),
                BASELINE_SALES_PLATFORM_CREATE_DATE,
            )
            cursor.execute(STAGE_CREATE_SQL)
            conn.commit()
            stage_batch: list[tuple[int, tuple[object, ...]]] = []
            for excel_row in excel_rows:
                data_rows += 1
                source_row = data_rows + 1
                db_row = {
                    column: convert(
                        column,
                        excel_row.get(header),
                        source_row=source_row,
                        source_name=source_name,
                    )
                    for header, column in IMPORT_COLUMN_MAP
                }
                _validate_erp_row(
                    db_row,
                    source_row,
                    require_allocation_fields=require_allocation_fields,
                )
                row_create_date = str(db_row["create_date"])
                if create_date is None:
                    create_date = row_create_date
                elif row_create_date != create_date:
                    raise ERPImportError(
                        f"本次 ERP 数据包含多个快照日期: {create_date}, "
                        f"{row_create_date}（源数据第 {source_row} 行）"
                    )

                if apply_baseline_sales_platform(db_row, sales_platform_baseline):
                    reused_baseline_sales_platform += 1
                else:
                    new_sales_platform += 1

                if apply_sales_platform_system_engineer(
                    db_row,
                    system_engineer_mapping,
                ):
                    applied_system_engineer_mapping += 1
                else:
                    kept_excel_system_engineer += 1

                stage_batch.append(
                    (
                        source_row,
                        tuple(db_row[column] for column in IMPORT_COLUMNS),
                    )
                )
                if len(stage_batch) >= batch_size:
                    _insert_stage_batch(cursor, stage_batch)
                    conn.commit()
                    stage_batch.clear()
                    logger.info("已处理 %d 行 ...", data_rows)

            if stage_batch:
                _insert_stage_batch(cursor, stage_batch)
                conn.commit()
            if not data_rows or create_date is None:
                raise ERPImportError("ERP 数据源不包含可导入的数据行。")

            _validate_staged_snapshot(
                cursor,
                create_date=create_date,
                expected_rows=data_rows,
            )
            conn.begin()
            replaced_rows = _publish_staged_snapshot(
                cursor,
                create_date=create_date,
                expected_rows=data_rows,
            )
            conn.commit()
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        if isinstance(exc, ERPImportError):
            raise
        raise ERPImportError(f"ERP 快照导入失败: {exc}") from exc
    finally:
        conn.close()

    seconds = round(time.time() - started, 1)
    logger.info(
        "导入完成: 发布 %d, 替换旧快照 %d, 复用基准营销平台 %d, 套用体系工程师映射 %d, 耗时 %ss",
        data_rows,
        replaced_rows,
        reused_baseline_sales_platform,
        applied_system_engineer_mapping,
        seconds,
    )
    return {
        "file": source_name,
        "rows": data_rows,
        "inserted": data_rows,
        "updated": 0,
        "unchanged": 0,
        "skipped": 0,
        "published_rows": data_rows,
        "replaced_rows": replaced_rows,
        "reused_baseline_sales_platform": reused_baseline_sales_platform,
        "new_sales_platform": new_sales_platform,
        "applied_system_engineer_mapping": applied_system_engineer_mapping,
        "kept_excel_system_engineer": kept_excel_system_engineer,
        "create_dates": [create_date],
        "seconds": seconds,
    }


def import_erp_xlsx(config: MySQLConfig, file_path: Path, batch_size: int = 5000) -> dict:
    """Import a standard ERP workbook into MySQL."""
    logger.info("打开文件: %s", file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = find_standard_sheet(wb)
        headers = _header_labels(ws)

        def excel_rows() -> Iterable[dict[str, object]]:
            for row in ws.iter_rows(min_row=2, values_only=True):
                yield dict(zip(headers, row, strict=True))

        return _import_erp_records(
            config,
            excel_rows(),
            file_path.name,
            batch_size,
            require_allocation_fields=len(headers) == len(standard_headers()),
        )
    finally:
        wb.close()


def import_erp_dataframe(config: MySQLConfig, dataframe, batch_size: int = 5000) -> dict:
    """Import an in-memory standard ERP dataframe without writing Sheet1."""
    headers = standard_headers()
    if dataframe.columns.tolist() != headers:
        raise ValueError("内存 ERP 数据必须使用完整且有序的 78 列标准字段。")

    def dataframe_rows() -> Iterable[dict[str, object]]:
        for row in dataframe.itertuples(index=False, name=None):
            yield dict(zip(headers, row, strict=True))

    return _import_erp_records(
        config,
        dataframe_rows(),
        "<memory>",
        batch_size,
        require_allocation_fields=True,
    )


def main():
    parser = argparse.ArgumentParser(description="ERP Excel 数据入库")
    parser.add_argument("--file", required=True, help="ERP Excel 文件路径")
    parser.add_argument("--batch-size", type=int, default=5000, help="每批提交行数")
    args = parser.parse_args()

    from .config import load_settings
    settings = load_settings()
    result = import_erp_xlsx(settings.mysql, Path(args.file), args.batch_size)
    print(result)


if __name__ == "__main__":
    main()
