"""Auditable writeback helpers for sampled ticket compliance results."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SAMPLING_STATUS_FIELD_KEY = "field_1447"
FAILURE_REASON_FIELD_KEY = "field_166"
COMPLIANT_OPTION_VALUE = "4328151"
NONCOMPLIANT_OPTION_VALUE = "4328146"
UNSET_OPTION_VALUES = {"", "--", "4328175"}
HUBEI_PROVINCE = "湖北省"
HUBEI_COMPLIANCE_REPORT_TYPE = "hubei_compliance_check"


@dataclass(frozen=True)
class SamplingStatusSource:
    ticket_id: str
    compliant: bool
    failure_reason: str = ""


@dataclass(frozen=True)
class ReportScope:
    """Machine-readable scope embedded by the compliance report producer."""

    report_type: str
    province: str
    quality_period: str = "manual"
    start: datetime | None = None
    end: datetime | None = None


@dataclass(frozen=True)
class SamplingStatusPlan:
    ticket_id: str
    target_value: str
    action: str
    current_value: str | None


def load_sampling_status_sources(report_path: Path) -> list[SamplingStatusSource]:
    """Read ticket ids and conclusions from the report's ``工单详情`` sheet."""

    workbook = load_workbook(report_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["工单详情"]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("检查报告的工单详情页没有表头")
        indexes = {str(value).strip(): index for index, value in enumerate(headers) if value}
        try:
            ticket_index = indexes["工单ID"]
            compliant_index = indexes["是否合规"]
        except KeyError as exc:
            raise ValueError("检查报告缺少工单ID或是否合规列") from exc

        sources: list[SamplingStatusSource] = []
        for row in rows:
            ticket_id = row[ticket_index]
            conclusion = row[compliant_index]
            if ticket_id is None or conclusion is None:
                continue
            conclusion_text = str(conclusion).strip()
            if conclusion_text not in {"合规", "不合规"}:
                raise ValueError(f"工单 {ticket_id} 的是否合规值无法识别: {conclusion_text}")
            sources.append(
                SamplingStatusSource(
                    ticket_id=str(ticket_id).strip(),
                    compliant=conclusion_text == "合规",
                    failure_reason=_failure_reason_from_report_row(row, indexes, conclusion_text),
                )
            )
        return sources
    finally:
        workbook.close()


def load_report_scope(report_path: Path) -> ReportScope:
    """Read the producer-owned scope metadata required before any writeback."""

    workbook = load_workbook(report_path, read_only=True, data_only=True)
    try:
        try:
            worksheet = workbook["回写范围"]
        except KeyError as exc:
            raise ValueError("检查报告缺少回写范围，拒绝执行回写") from exc
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows or tuple(rows[0][:2]) != ("键", "值"):
            raise ValueError("检查报告的回写范围格式无效，拒绝执行回写")
        values = {
            str(row[0]).strip(): str(row[1]).strip()
            for row in rows[1:]
            if len(row) >= 2 and row[0] is not None and row[1] is not None
        }
        try:
            return ReportScope(
                report_type=values["report_type"],
                province=values["province"],
                quality_period=values.get("quality_period", "manual"),
                start=_optional_datetime(values.get("start")),
                end=_optional_datetime(values.get("end")),
            )
        except KeyError as exc:
            raise ValueError("检查报告的回写范围缺少报告类型或省份，拒绝执行回写") from exc
    finally:
        workbook.close()


def assert_hubei_writeback_scope(scope: ReportScope) -> None:
    """Reject every report except the dedicated Hubei compliance report."""

    if scope.report_type != HUBEI_COMPLIANCE_REPORT_TYPE:
        raise ValueError(f"不支持的检查报告类型: {scope.report_type}")
    if scope.province != HUBEI_PROVINCE:
        raise ValueError(f"回写仅允许湖北省报告，当前省份: {scope.province}")


def assert_monthly_overwrite_scope(scope: ReportScope) -> None:
    """Allow overwrites only for a producer-marked complete natural month."""

    start = scope.start
    end = scope.end
    if (
        scope.quality_period != "monthly"
        or start is None
        or end is None
        or start.day != 1
        or start.time() != datetime.min.time()
        or end.time() != datetime.min.time()
    ):
        raise ValueError("覆盖回写仅允许完整自然月的湖北月结报告")
    expected_end = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    if end != expected_end:
        raise ValueError("覆盖回写仅允许完整自然月的湖北月结报告")


def _optional_datetime(value: str | None) -> datetime | None:
    return datetime.fromisoformat(value) if value else None


def current_custom_field_value(detail: dict[str, Any] | None, field_key: str) -> str | None:
    """Return the raw current value for one custom field, if the API returned it."""

    if not isinstance(detail, dict):
        return None
    fields = detail.get("custom_fields")
    if not isinstance(fields, list):
        return None
    for field in fields:
        if not isinstance(field, dict):
            continue
        key = field.get("key") or field.get("field_key") or field.get("fieldKey")
        if str(key or "") != field_key:
            continue
        value = field.get("value")
        return None if value is None else str(value).strip()
    return None


def build_sampling_status_plan(
    sources: list[SamplingStatusSource],
    current_values: dict[str, str | None],
    *,
    overwrite_existing: bool = False,
) -> list[SamplingStatusPlan]:
    """Build status writes, optionally replacing a different existing conclusion."""

    plan: list[SamplingStatusPlan] = []
    for source in sources:
        target_value = COMPLIANT_OPTION_VALUE if source.compliant else NONCOMPLIANT_OPTION_VALUE
        current_value = current_values.get(source.ticket_id)
        if overwrite_existing:
            action = "skip_unchanged" if current_value == target_value else "update"
        else:
            action = "update" if (current_value or "") in UNSET_OPTION_VALUES else "skip_existing"
        plan.append(
            SamplingStatusPlan(
                ticket_id=source.ticket_id,
                target_value=target_value,
                action=action,
                current_value=current_value,
            )
        )
    return plan


def build_failure_reason_plan(
    sources: list[SamplingStatusSource],
    current_values: dict[str, str | None],
    *,
    overwrite_existing: bool = False,
) -> list[SamplingStatusPlan]:
    """Plan reason writes and, in overwrite mode, clear stale compliant reasons."""

    plan: list[SamplingStatusPlan] = []
    for source in sources:
        current_value = current_values.get(source.ticket_id)
        if overwrite_existing:
            if not source.compliant and not source.failure_reason:
                continue
            target_value = "" if source.compliant else source.failure_reason
            action = (
                "skip_unchanged"
                if (current_value or "").strip() == target_value.strip()
                else "update"
            )
        else:
            if source.compliant or not source.failure_reason:
                continue
            target_value = source.failure_reason
            action = "update" if not (current_value or "").strip() else "skip_existing"
        plan.append(
            SamplingStatusPlan(
                ticket_id=source.ticket_id,
                target_value=target_value,
                action=action,
                current_value=current_value,
            )
        )
    return plan


def _failure_reason_from_report_row(
    row: tuple[Any, ...], indexes: dict[str, int], conclusion: str
) -> str:
    if conclusion == "合规":
        return ""

    checks = (
        ("规则1-标题非空", "标题为空"),
        ("规则2-描述含操作/现象/日志", "问题描述缺少操作、现象或报错日志"),
        ("规则3-解决含原因/过程/脚本", "解决办法缺少原因、处理过程或附件"),
        ("规则4-非敷衍", "解决办法内容过于简单"),
    )
    reasons = [
        reason
        for header, reason in checks
        if header in indexes and not str(row[indexes[header]] or "").strip().startswith("✓")
    ]
    duplicate_index = indexes.get("是否重复")
    if duplicate_index is not None and str(row[duplicate_index] or "").strip():
        reasons.append("存在重复工单")
    return "；".join(reasons) or "未通过填写规范抽检"
